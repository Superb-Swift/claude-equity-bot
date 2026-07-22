#!/usr/bin/env python3
# =============================================================================
# t2_check.py  —  five-second T2 tripwire read (Q1-R re-open gate)
# =============================================================================
# PURPOSE
#   One command, every session, answers the only open T2 question:
#   are BOTH gates met right now? Reads the tracker Signals tab ONLY —
#   no weekly harness, no formula-cache dependency (returns are computed
#   from the Price columns, so a stale/empty Return-column cache is fine).
#
# THE GATE (from Phase3A_Closeout_Memo §5 / kickoff; ruling R2 pools eras):
#   T2 FIRES  <=>  depth: live 60-69 +5d-resolved n >= 88
#             AND  lead : non-compressed stratum (50-59 acc - 60-69 acc) <= -15pt
#   Both must hold simultaneously. If it fires -> run the Tier-2E diff.
#
# DEFINITIONS (identical to guardrail_trace.py --stratify-regime, hardcoded
# here so the reading never drifts):
#   live            = signals dated on/after LIVE_START (2026-06-15)
#   resolved (+5d)  = Price +5d cell is populated
#   Right?(5d)      = HOLD: |ret5| < 0.03 ; BUY: ret5 > 0 ; SELL: ret5 < 0
#   cohort regime   = cohort-wide +5d hit rate: <40% compressed, else non-comp
#   non-compressed  = mixed + clean cohorts (rate >= 40%)
#   ret5            = (Price+5d - Price@signal) / Price@signal   [computed]
#
# USAGE
#   python t2_check.py                                   # default tracker
#   python t2_check.py --tracker tracker_with_registry.xlsx
#   python t2_check.py --history                         # show the last N days' lead trail
#   exit code 0 = not fired (or armed-on-one); 10 = BOTH gates met (T2 fires)
# =============================================================================

import argparse, datetime as dt, sys
import os
import openpyxl

LIVE_START = dt.date(2026, 6, 15)
DEPTH_TARGET = 88
LEAD_TRIGGER = -15.0          # non-compressed lead <= this arms the lead gate
COMPRESSED_BELOW = 40.0       # cohort hit% < this = compressed (excluded)
COL = dict(date=1, ticker=2, signal=4, conf=5, dq=6, px_sig=7, px_5d=8)


def _d(v):
    return v.date() if isinstance(v, dt.datetime) else (v if isinstance(v, dt.date) else None)

def _num(v):
    return v if isinstance(v, (int, float)) else None


def load_live_rows(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Signals"]
    rows, blank = [], 0
    for r in range(5, ws.max_row + 1):
        d = _d(ws.cell(r, COL["date"]).value)
        tk = ws.cell(r, COL["ticker"]).value
        if d is None and not tk:
            blank += 1
            if blank >= 30:
                break
            continue
        blank = 0
        if d is None or d < LIVE_START:
            continue
        conf = ws.cell(r, COL["conf"]).value
        conf = conf * 100 if isinstance(conf, (int, float)) and conf <= 1 else conf
        pxs, px5 = _num(ws.cell(r, COL["px_sig"]).value), _num(ws.cell(r, COL["px_5d"]).value)
        ret5 = (px5 - pxs) / pxs if (pxs and px5) else None
        rows.append(dict(date=d, conf=conf, sig=str(ws.cell(r, COL["signal"]).value or "").strip(),
                         ret5=ret5))
    wb.close()
    return rows


def right5(x):
    if x["ret5"] is None:
        return None
    s = x["sig"]
    return (abs(x["ret5"]) < 0.03 if s == "HOLD"
            else x["ret5"] > 0 if s == "BUY"
            else x["ret5"] < 0 if s == "SELL" else None)


def band_acc(rows, lo, hi):
    b = [x for x in rows if isinstance(x["conf"], (int, float)) and lo <= x["conf"] < hi
         and x["ret5"] is not None]
    h = sum(1 for x in b if right5(x))
    return len(b), (100 * h / len(b) if b else 0.0)


def compute(rows, as_of=None):
    if as_of:
        rows = [x for x in rows if x["date"] <= as_of]
    # depth: live 60-69 +5d-resolved
    depth_n, depth_acc = band_acc(rows, 60, 70)
    # non-compressed stratum by cohort regime
    noncomp = []
    for cd in sorted({x["date"] for x in rows}):
        c = [x for x in rows if x["date"] == cd and x["ret5"] is not None]
        if not c:
            continue
        rate = 100 * sum(1 for x in c if right5(x)) / len(c)
        if rate >= COMPRESSED_BELOW:
            noncomp.append(cd)
    nc = [x for x in rows if x["date"] in noncomp]
    n59, a59 = band_acc(nc, 50, 60)
    n69, a69 = band_acc(nc, 60, 70)
    lead = a59 - a69
    return dict(depth_n=depth_n, depth_acc=depth_acc, lead=lead,
                nc_cohorts=len(noncomp), n59=n59, a59=a59, n69=n69, a69=a69,
                last=max((x["date"] for x in rows), default=None))


def verdict(m):
    depth_ok = m["depth_n"] >= DEPTH_TARGET
    lead_ok = m["lead"] <= LEAD_TRIGGER
    return depth_ok, lead_ok, (depth_ok and lead_ok)


def fmt(m):
    depth_ok, lead_ok, fires = verdict(m)
    chk = lambda ok: "MET " if ok else "open"
    depth_gap = DEPTH_TARGET - m["depth_n"]
    lead_gap = m["lead"] - LEAD_TRIGGER          # >0 means above trigger (open)
    out = []
    out.append(f"T2 tripwire  (data through {m['last']})")
    out.append(f"  depth  n={m['depth_n']:>3} / {DEPTH_TARGET}   [{chk(depth_ok)}]"
               + (f"  {depth_gap} to go" if depth_gap > 0 else "  (crossed)"))
    out.append(f"  lead   {m['lead']:+5.1f}pt  (<= {LEAD_TRIGGER:.0f})   [{chk(lead_ok)}]"
               f"   {abs(lead_gap):.1f}pt {'above trigger' if lead_gap > 0 else 'past trigger'}")
    out.append(f"         non-comp {m['nc_cohorts']} cohorts | 50-59 {m['a59']:.1f}%(n={m['n59']})"
               f"  60-69 {m['a69']:.1f}%(n={m['n69']})")
    if fires:
        out.append("  -> *** T2 FIRES — both gates met. Run the Tier-2E diff. ***")
    else:
        # name the gate that is NOT satisfied (the one still to be met)
        if depth_ok and not lead_ok:
            why = "LEAD gate open (depth is met) — the lead is the deciding number"
        elif lead_ok and not depth_ok:
            why = "DEPTH gate open (lead is met) — depth is the deciding number"
        else:
            why = "NEITHER gate met"
        out.append(f"  -> NOT fired: {why}.")
    return "\n".join(out)


def main():
    ap = argparse.ArgumentParser(description="Five-second T2 tripwire read from the tracker.")
    ap.add_argument("--tracker",
                    default=os.path.join("logs", "claude_equity_bot_tracker.xlsx"),
                    help="the MASTER tracker (the one you paste into; its formula "
                         "cache is live). tracker_with_registry.xlsx is a GENERATED "
                         "report copy - never read it as input.")
    ap.add_argument("--history", action="store_true",
                    help="also print the lead/depth trail over the last 8 resolved days")
    args = ap.parse_args()

    rows = load_live_rows(args.tracker)
    m = compute(rows)
    print(fmt(m))

    if args.history:
        print("\n  lead trail (by as-of date):")
        days = sorted({x["date"] for x in rows if x["ret5"] is not None})[-8:]
        for cd in days:
            mm = compute(rows, as_of=cd)
            d_ok, l_ok, f = verdict(mm)
            flag = "FIRE" if f else ("depth" if d_ok else "----")
            print(f"    {cd}   depth {mm['depth_n']:>3}/{DEPTH_TARGET}   lead {mm['lead']:+5.1f}pt   {flag}")

    _, _, fires = verdict(m)
    sys.exit(10 if fires else 0)


if __name__ == "__main__":
    main()
