#!/usr/bin/env python3
"""
h1_lag_trace.py  —  Claude Equity Bot, Phase 3-A · H1 success metric

Re-runs the WMT-style confidence-update-lag trace from the Signals tab. This is
the success metric for H1 (prior-5-day price input): Phase 2 showed confidence
trailed price by ~3-5 sessions (the WMT arc held 62% through three ~-11% losses,
then caught up only as the losses shrank). H1 should compress that lag toward
0-1 sessions WITHOUT degrading the +5d band calibration.

The same script runs against the Phase 2 tracker (the baseline below) and, later,
the Phase 3-A tracker — so you can quantify the change.

WHAT IT PRODUCES
    1. Per-cohort trace for the ticker (default WMT): conf, trailing-5d move,
       forward +5d / +10d returns, +5d hit.
    2. A lag estimate: corr(confidence[t], trailing move[t-L]) scanned over
       L = 0..maxlag; the argmax L is the estimated update lag in sessions.
    3. The forward +10d arc (reproduces WMT -15.19 -> +4.31).
    4. A calibration guardrail: the cohort-window 3-band +5d hit rates
       (50-59% must keep leading).
    5. A SUCCESS-METRIC block: what 3-A must show vs this baseline.

USAGE
    python h1_lag_trace.py
    python h1_lag_trace.py --tracker claude_equity_bot_tracker.xlsx --ticker WMT
    python h1_lag_trace.py --ticker NVDA --window 5 --maxlag 5
"""

from __future__ import annotations
import argparse
import datetime as dt
import glob
import math
import os
import re
import sys

import openpyxl

SIGNALS_SHEET = "Signals"
FIRST_DATA_ROW = 5
LAST_ROW_CAP = 100_000
COL = dict(date=1, ticker=2, account=3, signal=4, conf=5, dq=6,
           p0=7, p5=8, p10=9, p20=10, r5=11, r10=12, r20=13, right5=14, notes=15)

COHORT_START = dt.date(2026, 5, 18)
COHORT_END = dt.date(2026, 5, 29)
# ---- optional date-window slice (--since / --until) --------------------------
# When either bound is set, the analysis AND the calibration guardrail are
# computed over [SINCE, UNTIL]. With no bounds: analysis uses all rows and the
# guardrail uses the certified Phase 2 cohort window (COHORT_START..COHORT_END).
SINCE = None
UNTIL = None


def _in_window(d):
    if not isinstance(d, dt.date):
        return False
    if SINCE is not None and d < SINCE:
        return False
    if UNTIL is not None and d > UNTIL:
        return False
    return True


def _guardrail_window():
    if SINCE is not None or UNTIL is not None:
        return (SINCE or dt.date.min, UNTIL or dt.date.max)
    return (COHORT_START, COHORT_END)


def _window_label():
    if SINCE is not None or UNTIL is not None:
        return (str(SINCE) if SINCE else 'earliest') + ' .. ' + (str(UNTIL) if UNTIL else 'latest')
    return 'ALL ROWS (guardrail = Phase 2 cohort window)'



# --------------------------------------------------------------------------- #
def as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    return v if isinstance(v, dt.date) else None


def conf_pct(v):
    if v is None:
        return None
    return v * 100.0 if v <= 1.0 else float(v)


def pearson(xs, ys):
    """Pearson r over paired, non-None numeric samples. None if < 3 pairs."""
    pairs = [(x, y) for x, y in zip(xs, ys) if x is not None and y is not None]
    n = len(pairs)
    if n < 3:
        return None
    mx = sum(p[0] for p in pairs) / n
    my = sum(p[1] for p in pairs) / n
    sxy = sum((p[0] - mx) * (p[1] - my) for p in pairs)
    sxx = sum((p[0] - mx) ** 2 for p in pairs)
    syy = sum((p[1] - my) ** 2 for p in pairs)
    if sxx == 0 or syy == 0:
        return None
    return sxy / math.sqrt(sxx * syy)


def pct(x, plus=False, dash="—"):
    if x is None:
        return dash
    s = f"{x*100:.2f}%"
    return f"+{s}" if (plus and x >= 0) else s


# --------------------------------------------------------------------------- #
def load_series(tracker, ticker):
    wb = openpyxl.load_workbook(tracker, data_only=True)
    ws = wb[SIGNALS_SHEET]
    rows = []
    blank = 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        d = as_date(ws.cell(row=r, column=COL["date"]).value)
        tk = ws.cell(row=r, column=COL["ticker"]).value
        if d is None and not tk:
            blank += 1
            if blank >= 25:
                break
            continue
        blank = 0
        if not _in_window(d) or str(tk).strip().upper() != ticker.upper():
            continue
        rows.append(dict(
            date=d,
            signal=str(ws.cell(row=r, column=COL["signal"]).value or "").strip(),
            conf=conf_pct(ws.cell(row=r, column=COL["conf"]).value),
            p0=ws.cell(row=r, column=COL["p0"]).value,
            r5=ws.cell(row=r, column=COL["r5"]).value,
            r10=ws.cell(row=r, column=COL["r10"]).value,
            right5=str(ws.cell(row=r, column=COL["right5"]).value or "").strip(),
        ))
    wb.close()
    rows.sort(key=lambda x: x["date"])
    return rows


def band_calibration(tracker):
    """Cohort-window 3-band +5d hit rates (the calibration guardrail)."""
    wb = openpyxl.load_workbook(tracker, data_only=True)
    ws = wb[SIGNALS_SHEET]
    agg = {"40-49": [0, 0], "50-59": [0, 0], "60-69": [0, 0]}
    glo, ghi = _guardrail_window()
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        d = as_date(ws.cell(row=r, column=COL["date"]).value)
        if d is None:
            continue
        if not (glo <= d <= ghi):
            continue
        p = conf_pct(ws.cell(row=r, column=COL["conf"]).value)
        if p is None:
            continue
        b = ("40-49" if 40 <= p < 50 else "50-59" if 50 <= p < 60
             else "60-69" if 60 <= p < 70 else None)
        if b is None:
            continue
        agg[b][1] += 1
        if str(ws.cell(row=r, column=COL["right5"]).value or "").strip().upper() == "YES":
            agg[b][0] += 1
    wb.close()
    return agg


# --------------------------------------------------------------------------- #
def load_raw_conf_map(patterns):
    """Parse Signal JSON log lines into {(date, TICKER): raw_conf_pct}.

    ANALYST NOTE (WS2 two-channel readout):
        Lever A logs both confidences on every line — "confidence" is the
        OPERATIVE (damped) value that flows to the tracker, and
        "confidence_raw" is the model's own output. The tracker therefore
        carries the operative channel; this loader recovers the RAW channel
        straight from the logs so the pre-registered lag test can be run on
        the model-side series (the honest Lever-B read). Date comes from the
        filename (signals_YYYY-MM-DD.log). PM diagnostic runs APPEND to the
        same signals_DATE.log as the AM run, so the FIRST occurrence per
        (date, ticker) wins — the AM-canonical read — and later PM lines
        can never shadow it.
    """
    tick_re = re.compile(r'"ticker":\s*"([A-Za-z.\-]+)"')
    raw_re = re.compile(r'"confidence_raw":\s*(\d+)')
    files = []
    for p in patterns:
        hits = sorted(glob.glob(p))
        files.extend(hits if hits else ([p] if os.path.exists(p) else []))
    m = {}
    for fp in files:
        stem = os.path.splitext(os.path.basename(fp))[0]
        try:
            d = dt.date.fromisoformat(stem.replace("signals_", ""))
        except ValueError:
            print(f"   [raw-from-logs] skip (no date in filename): {fp}")
            continue
        with open(fp, encoding="utf-8", errors="replace") as f:
            for line in f:
                if "Signal JSON" not in line:
                    continue
                tm, rm = tick_re.search(line), raw_re.search(line)
                if tm and rm:
                    key = (d, tm.group(1).upper())
                    if key not in m:  # FIRST occurrence wins - AM canonical
                        m[key] = float(rm.group(1))
    print(f"   [raw-from-logs] parsed {len(files)} file(s), "
          f"{len(m)} raw-confidence entrie(s)")
    return m


# --------------------------------------------------------------------------- #
def trace(tracker, ticker, window, maxlag, raw_map=None):
    rows = load_series(tracker, ticker)
    if not rows:
        sys.exit(f"ERROR: no rows for {ticker} in {tracker}")

    n_raw = 0
    if raw_map is not None:  # {} still prints the 0-override audit line
        for x in rows:
            v = raw_map.get((x["date"], ticker.upper()))
            if v is not None:
                x["conf"] = v
                n_raw += 1
        print(f"   [raw-from-logs] {n_raw} of {len(rows)} rows use log RAW "
              f"confidence; remainder keep tracker values (exact for "
              f"pre-deploy rows, where raw == operative by construction)")
    confs = [x["conf"] for x in rows]
    prices = [x["p0"] for x in rows]
    # trailing window-day move from the signal-capture price series
    trail = [None] * len(rows)
    for i in range(len(rows)):
        if i >= window and prices[i] is not None and prices[i - window]:
            trail[i] = prices[i] / prices[i - window] - 1.0

    # ---- per-cohort trace table ----
    print("\n" + "=" * 78)
    print(f"H1 LAG TRACE — {ticker}  (window={window}d, source: {os.path.basename(tracker)})")
    print("=" * 78)
    hdr = f"{'Date':<12}{'Sig':<6}{'Conf':>6}{'Trail'+str(window)+'d':>9}{'+5d':>9}{'+10d':>9}{'R5':>5}"
    print(hdr)
    print("-" * 78)
    for i, x in enumerate(rows):
        print(f"{x['date'].isoformat():<12}{x['signal']:<6}"
              f"{(str(int(x['conf']))+'%') if x['conf'] is not None else '—':>6}"
              f"{pct(trail[i]):>9}{pct(x['r5'], plus=True):>9}{pct(x['r10'], plus=True):>9}"
              f"{(x['right5'] or '—'):>5}")

    # ---- lag estimate: corr(conf[t], trail[t-L]) over L = 0..maxlag ----
    print("\nLAG SCAN — corr(confidence[t], trailing-move[t-L])  "
          "(positive = confidence tracks the move from L sessions ago):")
    scan = []
    for L in range(0, maxlag + 1):
        r = pearson(confs, trail) if L == 0 else pearson(confs[L:], trail[:-L])
        scan.append((L, r))
        print(f"   L={L} session(s): r = {r:+.3f}" if r is not None
              else f"   L={L} session(s): r = n/a (insufficient overlap)")
    # The lag is the L with the most POSITIVE correlation — i.e. how many
    # sessions late confidence catches up to the move. (argmax |r| would be
    # wrong: a NEGATIVE r at L=0 is the lag symptom, not contemporaneity.)
    pos = [(L, r) for L, r in scan if r is not None]
    best_L, best_r = (max(pos, key=lambda t: t[1]) if pos else (None, None))
    if best_L is not None:
        print(f"\n   >> Estimated update lag = {best_L} session(s)  "
              f"(peak positive r = {best_r:+.3f})")
        print("   0-1 = confidence tracks the move promptly; "
              "3-5 = the Phase 2 lag (catches up late).")
        if scan[0][1] is not None and scan[0][1] < 0:
            print(f"   (r at L=0 is {scan[0][1]:+.3f} — negative-at-0 flipping "
                  f"positive at L={best_L} is the lag signature: confidence sat "
                  "high through the recent adverse move.)")

    # ---- contemporaneous vs predictive correlations ----
    r_now = pearson(confs, trail)
    r_fwd = pearson(confs, [x["r5"] for x in rows])
    print(f"\n   corr(conf, trailing move)  = {r_now:+.3f}" if r_now is not None
          else "   corr(conf, trailing move)  = n/a")
    print(f"   corr(conf, forward +5d)    = {r_fwd:+.3f}" if r_fwd is not None
          else "   corr(conf, forward +5d)    = n/a")

    # ---- forward +10d arc ----
    arc = [x["r10"] for x in rows if x["r10"] is not None]
    if arc:
        arc_str = " -> ".join(f"{v*100:.2f}" for v in arc)
        print(f"\n+10d ARC ({ticker}): {arc_str}")
        print(f"   first {arc[0]*100:+.2f}%  ->  last {arc[-1]*100:+.2f}%  "
              f"({'self-corrected upward' if arc[-1] > arc[0] else 'no net recovery'})")

    # ---- calibration guardrail ----
    print("\nCALIBRATION GUARDRAIL — cohort-window +5d band hit rates "
          "(50-59% must lead):")
    agg = band_calibration(tracker)
    lead = max(agg, key=lambda b: (agg[b][0] / agg[b][1]) if agg[b][1] else -1)
    for b in ("40-49", "50-59", "60-69"):
        h, n = agg[b]
        rate = f"{h/n*100:.1f}%" if n else "—"
        mark = "  <-- leads" if b == lead else ""
        print(f"   {b}: {h}/{n} = {rate}{mark}")
    ok_cal = lead == "50-59"
    print(f"   [{'PASS' if ok_cal else 'WARN'}] 50-59% band "
          f"{'leads' if ok_cal else 'does NOT lead'}")

    # ---- success metric ----
    print("\n" + "-" * 78)
    print("SUCCESS METRIC (Phase 3-A vs this Phase 2 baseline)")
    print("-" * 78)
    print(f"   • Update lag shrinks from the Phase 2 ~3-5 sessions toward 0-1")
    print(f"     (this run's estimate: {best_L if best_L is not None else 'n/a'} session(s)).")
    print(f"   • corr(conf, trailing move) rises (confidence reacts to the recent")
    print(f"     move; this run: {r_now:+.3f})." if r_now is not None
          else "     move).")
    print(f"   • The 50-59% band keeps leading the +5d calibration "
          f"({'holds' if ok_cal else 'BROKEN'} here).")
    print(f"   • {ticker}-type names no longer hold high confidence through a")
    print(f"     multi-session adverse move.")
    print("\n   NOTE: per-ticker n is small (~19 cohorts) — read the trace table as")
    print("   the primary evidence and the correlations as directional diagnostics.")

    return dict(best_L=best_L, r_now=r_now, r_fwd=r_fwd, ok_cal=ok_cal,
                rows=rows, trail=trail, arc=arc, agg=agg)


def write_md(res, ticker, window, tracker, out_dir, suffix=""):
    rows, trail = res["rows"], res["trail"]
    lines = [f"# H1 Lag Trace — {ticker}",
             f"*Source: `{os.path.basename(tracker)}` · window {window}d · "
             f"generated {dt.datetime.now():%Y-%m-%d %H:%M}.*", "",
             f"**Estimated update lag: {res['best_L']} session(s)** · "
             f"corr(conf, trailing move) = {res['r_now']:+.3f} · "
             f"corr(conf, +5d) = {res['r_fwd']:+.3f}", "",
             f"| Date | Sig | Conf | Trail{window}d | +5d | +10d | R5 |",
             "| --- | --- | --- | --- | --- | --- | --- |"]
    for i, x in enumerate(rows):
        lines.append(
            f"| {x['date'].isoformat()} | {x['signal']} | "
            f"{(str(int(x['conf']))+'%') if x['conf'] is not None else '—'} | "
            f"{pct(trail[i])} | {pct(x['r5'], plus=True)} | "
            f"{pct(x['r10'], plus=True)} | {x['right5'] or '—'} |")
    if res["arc"]:
        lines += ["", f"**+10d arc:** {' → '.join(f'{v*100:.2f}' for v in res['arc'])} "
                      f"(first {res['arc'][0]*100:+.2f}% → last {res['arc'][-1]*100:+.2f}%)"]
    lines += ["", "## Calibration guardrail (cohort-window +5d)"]
    for b in ("40-49", "50-59", "60-69"):
        h, n = res["agg"][b]
        lines.append(f"- {b}: {h}/{n} = {h/n*100:.1f}%" if n else f"- {b}: —")
    path = os.path.join(out_dir, f"h1_lag_trace_{ticker}{suffix}.md")
    open(path, "w", encoding="utf-8").write("\n".join(lines) + "\n")
    return path


# --------------------------------------------------------------------------- #
def main(argv=None):
    ap = argparse.ArgumentParser(description="H1 confidence-lag trace / success metric.")
    ap.add_argument("--tracker", default="claude_equity_bot_tracker.xlsx")
    ap.add_argument("--ticker", default="WMT")
    ap.add_argument("--window", type=int, default=5, help="trailing-move window (sessions)")
    ap.add_argument("--maxlag", type=int, default=5, help="max lag to scan (sessions)")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--since", default=None,
                    help="only analyze rows on/after YYYY-MM-DD (e.g. 2026-06-15 = Phase 3-A only)")
    ap.add_argument("--until", default=None,
                    help="only analyze rows on/before YYYY-MM-DD")
    ap.add_argument("--raw-from-logs", nargs="*", default=None, metavar="LOG",
                    help="log files/globs; override tracker confidence with "
                         "the RAW (pre-damping) confidence parsed from Signal "
                         "JSON lines — the Lever-B model-side channel")
    ap.add_argument("--suffix", default="",
                    help="append to the output MD filename (e.g. _raw) so "
                         "raw-channel runs don't overwrite the operative MD")
    args = ap.parse_args(argv)
    global SINCE, UNTIL
    if args.since:
        SINCE = dt.date.fromisoformat(args.since)
    if args.until:
        UNTIL = dt.date.fromisoformat(args.until)
    if SINCE is not None or UNTIL is not None:
        print("")
        print("[DATE SLICE] analyzing " + _window_label())
        print("   (calibration guardrail uses this same window)")
    if not os.path.exists(args.tracker):
        sys.exit(f"ERROR: tracker not found: {args.tracker}")
    os.makedirs(args.out_dir, exist_ok=True)
    raw_map = (load_raw_conf_map(args.raw_from_logs)
               if args.raw_from_logs else None)
    res = trace(args.tracker, args.ticker, args.window, args.maxlag,
                raw_map=raw_map)
    md = write_md(res, args.ticker, args.window, args.tracker, args.out_dir,
                  suffix=args.suffix)
    print(f"\nWrote: {md}")


if __name__ == "__main__":
    main()
