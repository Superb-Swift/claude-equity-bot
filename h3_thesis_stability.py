#!/usr/bin/env python3
"""
h3_thesis_stability.py  —  Claude Equity Bot, Phase 3-A · H3 success metric

Measures THESIS STABILITY from the Signals tab. Phase 2's 5/29 AM/PM diagnostic
showed the bot was highly news-sensitive: 9 of 24 signals changed in six hours,
and confidence swung on sub-1% moves (VTI HOLD 52% -> BUY 68% on +0.10%; GLD
HOLD 42% -> 52% on +0.18%). H3 adds stable-thesis-vs-news-flow prompt language;
it should REDUCE that churn while preserving the +5d band calibration.

The intraday (AM/PM) signal needs dual runs. The Signals tab stores one row per
ticker per day (AM canonical), so this harness measures DAY-TO-DAY churn — the
same phenomenon at a longer timescale (analyst_notes, 5/29) — as the available
proxy, and reproduces a Phase 2 baseline the 3-A data can be compared against.

WHAT IT PRODUCES
    1. Per-ticker churn table: day-to-day signal flips, flip rate, mean |dConf|,
       and the small-move subset (|dPrice| < threshold) where the pathology lives.
    2. Portfolio aggregate: flip rate, mean |dConf|, and the small-move churn.
    3. The 5/29 intraday baseline for context.
    4. Calibration guardrail (50-59% +5d band must keep leading).
    5. A SUCCESS-METRIC block: what 3-A must show vs this baseline.

USAGE
    python h3_thesis_stability.py
    python h3_thesis_stability.py --tracker claude_equity_bot_tracker.xlsx
    python h3_thesis_stability.py --ticker NVDA --small-move 1.0
"""

from __future__ import annotations
import argparse
import datetime as dt
import os
import sys

import openpyxl

SIGNALS_SHEET = "Signals"
FIRST_DATA_ROW = 5
COL = dict(date=1, ticker=2, account=3, signal=4, conf=5, dq=6,
           p0=7, r5=11, right5=14)
COHORT_START = dt.date(2026, 5, 18)
COHORT_END = dt.date(2026, 5, 29)
# ---- optional date-window slice (--since / --until) --------------------------
# When either bound is set, the analysis AND the calibration guardrail are
# computed over [SINCE, UNTIL]. With no bounds: analysis uses all rows and the
# guardrail uses the certified Phase 2 cohort window (COHORT_START..COHORT_END).
SINCE = None
UNTIL = None


def _in_window(d):
    if not isinstance(d, dt.date):
        return False
    if SINCE is not None and d < SINCE:
        return False
    if UNTIL is not None and d > UNTIL:
        return False
    return True


def _guardrail_window():
    if SINCE is not None or UNTIL is not None:
        return (SINCE or dt.date.min, UNTIL or dt.date.max)
    return (COHORT_START, COHORT_END)


def _window_label():
    if SINCE is not None or UNTIL is not None:
        return (str(SINCE) if SINCE else 'earliest') + ' .. ' + (str(UNTIL) if UNTIL else 'latest')
    return 'ALL ROWS (guardrail = Phase 2 cohort window)'



def as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    return v if isinstance(v, dt.date) else None


def conf_pct(v):
    if v is None:
        return None
    return v * 100.0 if v <= 1.0 else float(v)


def load_by_ticker(tracker):
    wb = openpyxl.load_workbook(tracker, data_only=True)
    ws = wb[SIGNALS_SHEET]
    by = {}
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        d = as_date(ws.cell(row=r, column=COL["date"]).value)
        tk = ws.cell(row=r, column=COL["ticker"]).value
        if not _in_window(d) or not tk:
            continue
        sig = str(ws.cell(row=r, column=COL["signal"]).value or "").strip().upper()
        if sig not in ("BUY", "SELL", "HOLD"):
            continue
        by.setdefault(str(tk).strip().upper(), []).append(dict(
            date=d, signal=sig,
            conf=conf_pct(ws.cell(row=r, column=COL["conf"]).value),
            p0=ws.cell(row=r, column=COL["p0"]).value,
        ))
    wb.close()
    for tk in by:
        by[tk].sort(key=lambda x: x["date"])
    return by


def band_calibration(tracker):
    wb = openpyxl.load_workbook(tracker, data_only=True)
    ws = wb[SIGNALS_SHEET]
    agg = {"40-49": [0, 0], "50-59": [0, 0], "60-69": [0, 0]}
    glo, ghi = _guardrail_window()
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        d = as_date(ws.cell(row=r, column=COL["date"]).value)
        if d is None or not (glo <= d <= ghi):
            continue
        p = conf_pct(ws.cell(row=r, column=COL["conf"]).value)
        if p is None:
            continue
        b = ("40-49" if 40 <= p < 50 else "50-59" if 50 <= p < 60
             else "60-69" if 60 <= p < 70 else None)
        if b is None:
            continue
        agg[b][1] += 1
        if str(ws.cell(row=r, column=COL["right5"]).value or "").strip().upper() == "YES":
            agg[b][0] += 1
    wb.close()
    return agg


def ticker_churn(rows, small_move):
    """Per-ticker day-to-day churn stats. small_move is a fraction (0.01 = 1%)."""
    steps = flips = 0
    dconf_all, dconf_small, flips_small, small_steps = [], [], 0, 0
    for i in range(1, len(rows)):
        a, b = rows[i - 1], rows[i]
        steps += 1
        flipped = a["signal"] != b["signal"]
        if flipped:
            flips += 1
        if a["conf"] is not None and b["conf"] is not None:
            dconf = abs(b["conf"] - a["conf"])
            dconf_all.append(dconf)
            if a["p0"] and b["p0"]:
                dpx = abs(b["p0"] / a["p0"] - 1.0)
                if dpx < small_move:
                    small_steps += 1
                    dconf_small.append(dconf)
                    if flipped:
                        flips_small += 1
    return dict(
        days=len(rows), steps=steps, flips=flips,
        flip_rate=(flips / steps) if steps else None,
        mean_dconf=(sum(dconf_all) / len(dconf_all)) if dconf_all else None,
        small_steps=small_steps, flips_small=flips_small,
        mean_dconf_small=(sum(dconf_small) / len(dconf_small)) if dconf_small else None,
    )


def fmt(x, suffix="", dash="—"):
    return dash if x is None else f"{x:.1f}{suffix}"


def run(tracker, ticker, small_move, out_dir):
    by = load_by_ticker(tracker)
    if not by:
        sys.exit(f"ERROR: no signals found in {tracker}")
    tickers = [ticker.upper()] if ticker else sorted(by)
    tickers = [t for t in tickers if t in by]
    if not tickers:
        sys.exit(f"ERROR: ticker not found: {ticker}")

    print("\n" + "=" * 80)
    scope = ticker.upper() if ticker else f"portfolio ({len(tickers)} tickers)"
    print(f"H3 THESIS-STABILITY TRACE — {scope}  "
          f"(small-move < {small_move*100:.1f}%, source: {os.path.basename(tracker)})")
    print("=" * 80)

    # ---- per-ticker table ----
    hdr = (f"{'Ticker':<7}{'days':>5}{'flips':>6}{'flip%':>7}"
           f"{'mean|dConf|':>12}{'sm-steps':>9}{'sm|dConf|':>10}{'sm-flips':>9}")
    print(hdr)
    print("-" * 80)
    agg_steps = agg_flips = agg_small = agg_small_flips = 0
    all_dconf, all_dconf_small = [], []
    rows_for_md = []
    for tk in tickers:
        c = ticker_churn(by[tk], small_move)
        agg_steps += c["steps"]; agg_flips += c["flips"]
        agg_small += c["small_steps"]; agg_small_flips += c["flips_small"]
        # rebuild raw lists for true aggregate means
        for i in range(1, len(by[tk])):
            a, b = by[tk][i - 1], by[tk][i]
            if a["conf"] is not None and b["conf"] is not None:
                dconf = abs(b["conf"] - a["conf"]); all_dconf.append(dconf)
                if a["p0"] and b["p0"] and abs(b["p0"] / a["p0"] - 1.0) < small_move:
                    all_dconf_small.append(dconf)
        print(f"{tk:<7}{c['days']:>5}{c['flips']:>6}"
              f"{fmt((c['flip_rate'] or 0)*100,'%' ):>7}"
              f"{fmt(c['mean_dconf']):>12}{c['small_steps']:>9}"
              f"{fmt(c['mean_dconf_small']):>10}{c['flips_small']:>9}")
        rows_for_md.append((tk, c))

    flip_rate = (agg_flips / agg_steps) if agg_steps else None
    mean_dconf = (sum(all_dconf) / len(all_dconf)) if all_dconf else None
    mean_dconf_small = (sum(all_dconf_small) / len(all_dconf_small)) if all_dconf_small else None
    small_flip_rate = (agg_small_flips / agg_small) if agg_small else None

    # ---- portfolio aggregate ----
    print("\nPORTFOLIO AGGREGATE (Phase 2 baseline):")
    print(f"   day-to-day signal flips : {agg_flips} / {agg_steps} steps "
          f"= {fmt((flip_rate or 0)*100, '%')}")
    print(f"   mean |dConf| per step   : {fmt(mean_dconf)} pts")
    print(f"   small moves (<{small_move*100:.0f}%) : {agg_small} steps; "
          f"mean |dConf| {fmt(mean_dconf_small)} pts; "
          f"{agg_small_flips} flips ({fmt((small_flip_rate or 0)*100, '%')})  "
          f"<-- the H3 pathology")

    # ---- intraday baseline ----
    print("\nINTRADAY BASELINE (5/29 AM/PM diagnostic, the original H3 evidence):")
    print("   9 of 24 signals changed in ~6h; e.g. VTI HOLD 52% -> BUY 68% on +0.10%,")
    print("   GLD HOLD 42% -> 52% on +0.18%. True intraday measurement needs AM/PM")
    print("   dual-runs (logs/diagnostic_2026-05-29_PM.txt); the day-to-day numbers")
    print("   above are the tracker-available proxy.")

    # ---- calibration guardrail ----
    agg = band_calibration(tracker)
    lead = max(agg, key=lambda b: (agg[b][0] / agg[b][1]) if agg[b][1] else -1)
    print("\nCALIBRATION GUARDRAIL — cohort-window +5d band hit rates "
          "(50-59% must lead):")
    for b in ("40-49", "50-59", "60-69"):
        h, n = agg[b]
        print(f"   {b}: {h}/{n} = {h/n*100:.1f}%" + ("  <-- leads" if b == lead else "")
              if n else f"   {b}: —")
    ok_cal = lead == "50-59"
    print(f"   [{'PASS' if ok_cal else 'WARN'}] 50-59% band "
          f"{'leads' if ok_cal else 'does NOT lead'}")

    # ---- success metric ----
    print("\n" + "-" * 80)
    print("SUCCESS METRIC (Phase 3-A vs this Phase 2 baseline)")
    print("-" * 80)
    print(f"   • Day-to-day flip rate falls below {fmt((flip_rate or 0)*100, '%')}.")
    print(f"   • Mean |dConf| on small (<{small_move*100:.0f}%) moves falls below "
          f"{fmt(mean_dconf_small)} pts — the model stops re-rating on noise.")
    print(f"   • Signal flips on small moves ({agg_small_flips} here) trend toward 0.")
    print(f"   • The 50-59% +5d band keeps leading ({'holds' if ok_cal else 'BROKEN'}).")
    print("\n   NOTE: read this as day-to-day stability; pair with an AM/PM dual-run")
    print("   on a news-heavy day to confirm the intraday effect directly.")

    # ---- markdown ----
    md = [f"# H3 Thesis-Stability Trace — {scope}",
          f"*Source: `{os.path.basename(tracker)}` · small-move < {small_move*100:.1f}% · "
          f"generated {dt.datetime.now():%Y-%m-%d %H:%M}.*", "",
          f"**Portfolio: {agg_flips}/{agg_steps} flips = {fmt((flip_rate or 0)*100,'%')} · "
          f"mean |dConf| {fmt(mean_dconf)} pts · "
          f"small-move |dConf| {fmt(mean_dconf_small)} pts · "
          f"{agg_small_flips} flips on small moves**", "",
          "| Ticker | days | flips | flip% | mean\\|dConf\\| | sm-steps | sm\\|dConf\\| | sm-flips |",
          "| --- | --- | --- | --- | --- | --- | --- | --- |"]
    for tk, c in rows_for_md:
        md.append(f"| {tk} | {c['days']} | {c['flips']} | "
                  f"{fmt((c['flip_rate'] or 0)*100,'%')} | {fmt(c['mean_dconf'])} | "
                  f"{c['small_steps']} | {fmt(c['mean_dconf_small'])} | {c['flips_small']} |")
    md += ["", "## Calibration guardrail (cohort-window +5d)"]
    for b in ("40-49", "50-59", "60-69"):
        h, n = agg[b]
        md.append(f"- {b}: {h}/{n} = {h/n*100:.1f}%" if n else f"- {b}: —")
    suffix = ticker.upper() if ticker else "portfolio"
    path = os.path.join(out_dir, f"h3_thesis_stability_{suffix}.md")
    open(path, "w", encoding="utf-8").write("\n".join(md) + "\n")
    print(f"\nWrote: {path}")


def main(argv=None):
    ap = argparse.ArgumentParser(description="H3 thesis-stability / churn success metric.")
    ap.add_argument("--tracker", default="claude_equity_bot_tracker.xlsx")
    ap.add_argument("--ticker", default=None, help="single ticker (default: whole portfolio)")
    ap.add_argument("--small-move", type=float, default=1.0,
                    help="small-move threshold in PERCENT (default 1.0)")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--since", default=None,
                    help="only analyze rows on/after YYYY-MM-DD (e.g. 2026-06-15 = Phase 3-A only)")
    ap.add_argument("--until", default=None,
                    help="only analyze rows on/before YYYY-MM-DD")
    args = ap.parse_args(argv)
    global SINCE, UNTIL
    if args.since:
        SINCE = dt.date.fromisoformat(args.since)
    if args.until:
        UNTIL = dt.date.fromisoformat(args.until)
    if SINCE is not None or UNTIL is not None:
        print("")
        print("[DATE SLICE] analyzing " + _window_label())
        print("   (calibration guardrail uses this same window)")
    if not os.path.exists(args.tracker):
        sys.exit(f"ERROR: tracker not found: {args.tracker}")
    os.makedirs(args.out_dir, exist_ok=True)
    run(args.tracker, args.ticker, args.small_move / 100.0, args.out_dir)


if __name__ == "__main__":
    main()
