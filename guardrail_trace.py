#!/usr/bin/env python3
"""
guardrail_trace.py — Confidence-band guardrail compression trace.

The Phase 2 GO/NO-GO guardrail is "the 50-59% confidence band leads +5d accuracy."
Certified snapshot was 61.3% (n=216). As live +5d outcomes resolve, the cumulative
50-59% rate has eased (61.3 -> ~58.7). This harness traces, cohort by cohort, the
running-cumulative +5d hit rate of each confidence band so we can see whether the
50-59% lead over 60-69% is leveling off or trending toward a crossover (the failure
condition for the guardrail).

Source of truth: the Signals tab. A signal counts once its +5d outcome (Was-Right-5d,
col N) has resolved. Bands use closing-price-based Was-Right exactly as the tracker
computes it. Read-only on the source; never transcribes.

With --embed-into, also writes a "Live Cohort Scoreboard" sheet: one row per Phase 3-A
cohort (date >= --phase3a-start) with the per-cohort pairwise verdict (50-59% > 60-69%),
the cumulative lead through that cohort, and a regime flag from the cohort-wide hit rate.

Usage:
    python3 guardrail_trace.py --tracker claude_equity_bot_tracker.xlsx
    python3 guardrail_trace.py --tracker t.xlsx --since 2026-05-18 --until 2026-06-11
    python3 guardrail_trace.py --tracker t.xlsx --out-dir viz   # PNG destination
"""
from __future__ import annotations
import argparse
import datetime as dt
import os
import sys
from collections import defaultdict

import openpyxl

SIGNALS_SHEET = "Signals"
FIRST_DATA_ROW = 5
COL = dict(date=1, conf=5, right5=14)

# Bands as [low, high) on the confidence FRACTION (0.xx). 70%+ is the actionable
# tier (tiny n in Phase 2) and is reported separately if present.
BANDS = [("40-49%", 0.40, 0.50), ("50-59%", 0.50, 0.60), ("60-69%", 0.60, 0.70)]
LEAD_BAND = "50-59%"      # the guardrail band
RIVAL_BAND = "60-69%"     # the band it must stay above
CERTIFIED_LEAD = 0.613    # Phase 2 certified snapshot for 50-59% (n=216)
SHEET_NAME = "Guardrail Trace"   # name of the sheet written by --embed-into

# --- Live Cohort Scoreboard (Phase 3-A) ---
PHASE3A_START = dt.date(2026, 6, 15)        # first live advisory cohort
SCOREBOARD_SHEET = "Live Cohort Scoreboard"
# Regime flag keyed off the cohort-wide +5d hit rate (all confidences pooled).
# When forward returns are broadly adverse, every band compresses toward ~50% and
# the band ordering carries little information — so low-hit cohorts get flagged and
# weighted lightly. These thresholds are reporting aids, not statistical cutoffs.
REGIME_COMPRESSED = 0.40   # below: broad misses, bands can't separate
REGIME_CLEAN = 0.55        # above: bands have room to rank cleanly


def conf_frac(v):
    if v is None:
        return None
    return float(v) / 100.0 if v > 1.0 else float(v)


def band_of(c):
    if c is None:
        return None
    for name, lo, hi in BANDS:
        if lo <= c < hi:
            return name
    return None


def resolved(v):
    return str(v).strip().upper() in ("YES", "NO")


def is_hit(v):
    return str(v).strip().upper() == "YES"


def regime_of(cohort_hit):
    """Map a cohort-wide +5d hit rate to its regime label. Single source of truth
    for both the scoreboard's Regime column and the --stratify-regime mode."""
    if cohort_hit is None:
        return "\u2013"   # en-dash: no resolved signals
    if cohort_hit < REGIME_COMPRESSED:
        return "compressed"
    if cohort_hit > REGIME_CLEAN:
        return "clean"
    return "mixed"


def load_cohorts(tracker, since, until):
    """Return ordered list of (date, {band: [hits, n]}) for resolved +5d signals."""
    wb = openpyxl.load_workbook(tracker, data_only=True, read_only=True)
    ws = wb[SIGNALS_SHEET]
    # per (date, band) tally
    agg = defaultdict(lambda: defaultdict(lambda: [0, 0]))  # date -> band -> [hits, n]
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW):
        d = row[COL["date"] - 1].value
        if d is None:
            continue
        dd = d.date() if isinstance(d, dt.datetime) else d
        if not isinstance(dd, dt.date):
            continue
        if since and dd < since:
            continue
        if until and dd > until:
            continue
        r5 = row[COL["right5"] - 1].value
        if not resolved(r5):
            continue
        b = band_of(conf_frac(row[COL["conf"] - 1].value))
        if b is None:
            continue
        cell = agg[dd][b]
        cell[1] += 1
        if is_hit(r5):
            cell[0] += 1
    wb.close()
    return sorted(agg.items())


def load_cohort_totals(tracker, since, until):
    """Return {date: [hits, n]} over ALL resolved +5d signals in a cohort,
    regardless of confidence band (incl. <40% and 70%+). This pooled rate is the
    cohort-wide difficulty signal the regime flag keys off."""
    wb = openpyxl.load_workbook(tracker, data_only=True, read_only=True)
    ws = wb[SIGNALS_SHEET]
    totals = defaultdict(lambda: [0, 0])  # date -> [hits, n]
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW):
        d = row[COL["date"] - 1].value
        if d is None:
            continue
        dd = d.date() if isinstance(d, dt.datetime) else d
        if not isinstance(dd, dt.date):
            continue
        if since and dd < since:
            continue
        if until and dd > until:
            continue
        r5 = row[COL["right5"] - 1].value
        if not resolved(r5):
            continue
        totals[dd][1] += 1
        if is_hit(r5):
            totals[dd][0] += 1
    wb.close()
    return totals


def trace(cohorts):
    """Build per-cohort + cumulative rows. Returns (table_rows, series)."""
    cum = defaultdict(lambda: [0, 0])  # band -> [hits, n]
    table, series = [], defaultdict(list)
    dates = []
    for d, bands in cohorts:
        for name, _, _ in BANDS:
            h, n = bands.get(name, [0, 0])
            cum[name][0] += h
            cum[name][1] += n
        dates.append(d)
        row = {"date": d}
        for name, _, _ in BANDS:
            ch, cn = bands.get(name, [0, 0])
            tot_h, tot_n = cum[name]
            row[name] = (ch, cn, (tot_h / tot_n if tot_n else None))
            series[name].append(tot_h / tot_n if tot_n else None)
        table.append(row)
    return table, series, dates


def print_table(table):
    L, R = LEAD_BAND, RIVAL_BAND
    print(f"\n{'cohort':<11}  {'50-59 (coh)':>11}  {'60-69 (coh)':>11}  "
          f"{'cum 50-59':>9}  {'cum 60-69':>9}  {'cum 40-49':>9}  {'lead gap':>8}")
    print("-" * 86)
    for r in table:
        lh, ln, lc = r[L]
        rh, rn, rc = r[R]
        _, _, oc = r["40-49%"]
        coh_l = f"{lh}/{ln}" if ln else "  –"
        coh_r = f"{rh}/{rn}" if rn else "  –"
        gap = (lc - rc) if (lc is not None and rc is not None) else None
        print(f"{r['date']:%Y-%m-%d}  {coh_l:>11}  {coh_r:>11}  "
              f"{(f'{lc:.1%}' if lc is not None else '–'):>9}  "
              f"{(f'{rc:.1%}' if rc is not None else '–'):>9}  "
              f"{(f'{oc:.1%}' if oc is not None else '–'):>9}  "
              f"{(f'{gap*100:+.1f}pt' if gap is not None else '–'):>8}")


def make_chart_weekly(cohorts, series, dates, out_path):
    """Weekly-aggregated guardrail chart (fixes the unbounded-x squish).

    WHY: one x-tick per cohort grows +1 every trading day (39 -> 49 -> ...),
    so the axis always eventually squishes. Aggregating cohorts into ISO weeks
    changes growth to ~+1 tick/week and buries less signal in noise (per-cohort
    bands swing 33-88% on n=5-11; weekly they read on n=18-57).

    Bars   = that week's own hit rate per band (recent behaviour).
    Lines  = cumulative hit rate (the long-run trend the trace has always shown).
    """
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter
    from collections import OrderedDict

    # aggregate per-cohort hits/n into ISO weeks
    wk = OrderedDict()
    for (d, bands) in cohorts:
        key = f"wk{d.isocalendar()[1]}"
        cell = wk.setdefault(key, {n: [0, 0] for n, _, _ in BANDS})
        for name, _, _ in BANDS:
            h, n = bands.get(name, [0, 0])
            cell[name][0] += h
            cell[name][1] += n
    weeks = list(wk.keys())
    x = list(range(len(weeks)))

    def wrate(name):
        return [(wk[w][name][0] / wk[w][name][1] if wk[w][name][1] else None)
                for w in weeks]

    # cumulative value at each week's LAST cohort (reuse the existing series)
    last_idx, seen = [], set()
    for i, d in enumerate(dates):
        seen.add(f"wk{d.isocalendar()[1]}")
        if len(seen) == len(last_idx) + 1:
            last_idx.append(i)
        else:
            last_idx[-1] = i
    cum_lead = [series[LEAD_BAND][i] for i in last_idx]
    cum_rival = [series[RIVAL_BAND][i] for i in last_idx]

    fig, ax = plt.subplots(figsize=(11, 6))
    w = 0.36
    lead_w = wrate(LEAD_BAND)
    rival_w = wrate(RIVAL_BAND)
    ax.bar([i - w / 2 for i in x], [v if v is not None else 0 for v in lead_w],
           width=w, color="#1a73e8", alpha=0.30, label="50-59% (that week)", zorder=2)
    ax.bar([i + w / 2 for i in x], [v if v is not None else 0 for v in rival_w],
           width=w, color="#d93025", alpha=0.30, label="60-69% (that week)", zorder=2)
    ax.plot(x, cum_lead, "-o", color="#1a73e8", lw=2.4, ms=5,
            label="50-59% (cumulative)", zorder=4)
    ax.plot(x, cum_rival, "-o", color="#d93025", lw=2.0, ms=4,
            label="60-69% (cumulative)", zorder=3)

    ax.axhline(0.50, ls=":", lw=1, color="#9aa0a6", zorder=1)
    ax.axhline(CERTIFIED_LEAD, ls="--", lw=1, color="#1a73e8", alpha=0.45, zorder=1)
    ax.text(x[0], CERTIFIED_LEAD + 0.006, "50-59 certified snapshot 61.3%",
            color="#1a73e8", fontsize=8, ha="left", va="bottom")

    lead_gap = (cum_lead[-1] - cum_rival[-1]) * 100 if cum_lead and cum_rival else 0.0
    ax.set_title("Guardrail trace (weekly) — bars: that week · lines: cumulative"
                 f"   |   final lead {lead_gap:+.1f}pt",
                 fontsize=13, fontweight="bold", pad=12)
    ax.set_xlabel("ISO week (cohorts aggregated)", fontsize=10)
    ax.set_ylabel("+5d hit rate", fontsize=10)
    ax.set_xticks(x)
    ax.set_xticklabels(weeks, fontsize=9)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"{v:.0%}"))
    ax.set_ylim(0.0, 1.0)
    ax.grid(axis="y", color="#eee", lw=0.8)
    for sp in ("top", "right"):
        ax.spines[sp].set_visible(False)
    ax.legend(loc="lower left", frameon=False, fontsize=8, ncol=2)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    return weeks


def make_chart(table, series, dates, out_path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter

    x = list(range(len(dates)))
    labels = [f"{d.month}/{d.day}" for d in dates]
    fig, ax = plt.subplots(figsize=(11, 6))

    ax.axhline(0.50, ls=":", lw=1, color="#9aa0a6", zorder=1)
    ax.text(x[-1], 0.505, "coin flip (50%)", color="#80868b", fontsize=8,
            ha="right", va="bottom")
    ax.axhline(CERTIFIED_LEAD, ls="--", lw=1, color="#1a73e8", alpha=0.45, zorder=1)
    ax.text(x[0], CERTIFIED_LEAD + 0.004, "50-59 certified snapshot 61.3%",
            color="#1a73e8", fontsize=8, ha="left", va="bottom")

    ax.plot(x, series[LEAD_BAND], "-o", color="#1a73e8", lw=2.4, ms=5,
            label="50-59% (guardrail band)", zorder=4)
    ax.plot(x, series[RIVAL_BAND], "-o", color="#d93025", lw=2.0, ms=4,
            label="60-69%", zorder=3)
    ax.plot(x, series["40-49%"], "-o", color="#9aa0a6", lw=1.3, ms=3,
            alpha=0.8, label="40-49%", zorder=2)

    # shade the lead gap at the final point
    yl, yr = series[LEAD_BAND][-1], series[RIVAL_BAND][-1]
    ax.annotate(f"lead {((yl - yr) * 100):+.1f}pt",
                xy=(x[-1], (yl + yr) / 2), xytext=(x[-1] - 1.6, (yl + yr) / 2),
                fontsize=9, color="#202124", va="center",
                arrowprops=dict(arrowstyle="-", color="#bdc1c6", lw=0.8))

    ax.set_title("Guardrail compression — cumulative +5d hit rate by confidence band",
                 fontsize=13, fontweight="bold", pad=12)
    ax.set_xlabel("cohort (cumulative through date)", fontsize=10)
    ax.set_ylabel("cumulative +5d hit rate", fontsize=10)
    ax.set_xticks(x)
    ax.set_xticklabels(labels, rotation=0, fontsize=8)
    ax.yaxis.set_major_formatter(FuncFormatter(lambda v, _: f"{v:.0%}"))
    ax.set_ylim(0.35, 0.70)
    ax.grid(axis="y", color="#eee", lw=0.8)
    for s in ("top", "right"):
        ax.spines[s].set_visible(False)
    ax.legend(loc="lower left", frameon=False, fontsize=9)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def embed_into_workbook(table, series, dates, png_path, embed_path):
    """Write/refresh a 'Guardrail Trace' sheet — auditable data table + the
    trajectory chart image + a summary — into embed_path so the chart travels
    with the workbook. Idempotent: replaces the sheet if it already exists."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.drawing.image import Image as XLImage

    wb = openpyxl.load_workbook(embed_path)
    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME)

    hdr_fill = PatternFill("solid", fgColor="1A237E")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "Guardrail Compression Trace — cumulative +5d hit rate by confidence band"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (f"Auto-generated by guardrail_trace.py · source: Signals tab · "
                f"as of {dates[-1]:%Y-%m-%d}. The 50-59% band must lead 60-69%.")
    ws["A2"].font = Font(italic=True, size=9, color="666666")

    headers = ["Cohort", "50-59 (coh)", "60-69 (coh)", "cum 50-59",
               "cum 60-69", "cum 40-49", "lead gap (pt)"]
    hr = 4
    for j, h in enumerate(headers, start=1):
        c = ws.cell(hr, j, h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")
        c.border = bd

    L, R = LEAD_BAND, RIVAL_BAND
    for i, r in enumerate(table):
        rr = hr + 1 + i
        lh, ln, lc = r[L]
        rh, rn, rc = r[R]
        _, _, oc = r["40-49%"]
        ws.cell(rr, 1, r["date"]).number_format = "m/d"
        ws.cell(rr, 2, (f"{lh}/{ln}" if ln else "–")).alignment = Alignment(horizontal="center")
        ws.cell(rr, 3, (f"{rh}/{rn}" if rn else "–")).alignment = Alignment(horizontal="center")
        for col, val in ((4, lc), (5, rc), (6, oc)):
            ws.cell(rr, col, val).number_format = "0.0%"
        gap = (lc - rc) if (lc is not None and rc is not None) else None
        ws.cell(rr, 7, (round(gap * 100, 1) if gap is not None else None)).number_format = "0.0"
        for j in range(1, 8):
            ws.cell(rr, j).border = bd
    last = hr + len(table)
    for j, w in enumerate([9, 11, 11, 10, 10, 10, 12], start=1):
        ws.column_dimensions[chr(64 + j)].width = w

    # summary block under the table
    lc, rc, oc = series[L][-1], series[R][-1], series["40-49%"][-1]
    gaps = [series[L][i] - series[R][i] for i in range(len(dates))
            if series[L][i] is not None and series[R][i] is not None]
    sr = last + 2
    ws.cell(sr, 1, f"As of {dates[-1]:%Y-%m-%d}").font = Font(bold=True)
    ws.cell(sr + 1, 1, f"50-59% {lc:.1%}   ·   60-69% {rc:.1%}   ·   40-49% {oc:.1%}")
    ws.cell(sr + 2, 1, f"lead (50-59 over 60-69): {(lc - rc) * 100:+.1f}pt"
                       f"    ·    drift from certified 61.3%: {(lc - CERTIFIED_LEAD) * 100:+.1f}pt")
    if len(gaps) >= 4:
        recent = gaps[-3:]
        slope = recent[-1] - recent[0]
        trend = ("widening" if slope > 0.01 else
                 "narrowing" if slope < -0.01 else "roughly flat")
        ws.cell(sr + 3, 1, f"lead gap, last 3 cohorts: {trend} "
                           f"({recent[0] * 100:+.1f}pt → {recent[-1] * 100:+.1f}pt)")

    # embed the polished trajectory chart to the right of the table
    if png_path and os.path.exists(png_path):
        img = XLImage(png_path)
        scale = min(1.0, 760.0 / float(img.width))   # cap width ~760 px
        img.width = int(img.width * scale)
        img.height = int(img.height * scale)
        ws.add_image(img, "I4")

    wb.save(embed_path)


def build_scoreboard(cohorts, totals, series, dates, phase3a_start):
    """One row per Phase 3-A cohort (date >= phase3a_start). Each row carries the
    per-cohort guardrail verdict, the cumulative lead through that cohort, and a
    regime flag from the cohort-wide hit rate.

    'Held' is the guardrail's PAIRWISE claim (50-59% > 60-69%) — deliberately
    distinct from the tracker's stricter 'is 50-59 the top band' Pattern, so a
    cohort can read NO in the tracker yet still hold the pairwise claim."""
    L, R = LEAD_BAND, RIVAL_BAND
    date_index = {d: i for i, d in enumerate(dates)}
    rows = []
    for d, bands in cohorts:
        if d < phase3a_start:
            continue
        th, tn = totals.get(d, [0, 0])
        cohort_hit = (th / tn) if tn else None
        rates = {}
        for name, _, _ in BANDS:
            h, n = bands.get(name, [0, 0])
            rates[name] = (h, n, (h / n) if n else None)
        l_h, l_n, l_rate = rates[L]
        _, _, r_rate = rates[R]
        if l_rate is not None and r_rate is not None:
            held = "YES" if l_rate > r_rate else ("TIE" if l_rate == r_rate else "NO")
        else:
            held = "n/a"
        present = [(name.replace("%", ""), rates[name][2], rates[name][1])
                   for name, _, _ in BANDS if rates[name][1] > 0]
        present.sort(key=lambda x: x[1], reverse=True)
        parts = []
        for k, (nm, rt, nn) in enumerate(present):
            sep = "" if k == 0 else (" = " if abs(rt - present[k - 1][1]) < 1e-9 else " > ")
            parts.append(f"{sep}{nm} {rt:.0%}({nn})")
        order = "".join(parts) if present else "–"
        idx = date_index[d]
        cl, cr = series[L][idx], series[R][idx]
        cum_lead = (cl - cr) if (cl is not None and cr is not None) else None
        regime = regime_of(cohort_hit)
        rows.append(dict(date=d, n=tn, l_h=l_h, l_n=l_n, l_rate=l_rate, held=held,
                         order=order, cum_lead=cum_lead, cohort_hit=cohort_hit,
                         regime=regime))
    return rows


def print_scoreboard(rows):
    print("\n" + "=" * 60)
    print("LIVE COHORT SCOREBOARD (Phase 3-A) — 50-59% vs 60-69% out-of-sample")
    if not rows:
        print("  (no Phase 3-A cohorts have resolved their +5d yet)")
        return
    print(f"{'cohort':<11}{'n':>4}  {'50-59%':>13}  {'held':>4}  "
          f"{'band order (by hit)':<26}{'cum lead':>9}  {'coh hit':>7}  {'regime':<11}")
    print("-" * 98)
    for r in rows:
        rate = (f"{r['l_h']} of {r['l_n']} ({r['l_rate']:.0%})" if r['l_n'] else "–")
        cl = (f"{r['cum_lead'] * 100:+.1f}pt" if r['cum_lead'] is not None else "–")
        ch = (f"{r['cohort_hit']:.0%}" if r['cohort_hit'] is not None else "–")
        print(f"{r['date']:%Y-%m-%d}{r['n']:>4}  {rate:>13}  {r['held']:>4}  "
              f"{r['order']:<26}{cl:>9}  {ch:>7}  {r['regime']:<11}")
    held_yes = sum(1 for r in rows if r["held"] == "YES")
    held_tot = sum(1 for r in rows if r["held"] in ("YES", "NO"))
    print(f"\n  pairwise held (50-59 > 60-69): {held_yes} of {held_tot} live cohorts"
          "   ·   compressed cohorts are low-information")


def embed_scoreboard_sheet(rows, embed_path, as_of):
    """Write/refresh a 'Live Cohort Scoreboard' sheet into embed_path. Idempotent:
    replaces the sheet if present. No formulas — static, recalc-safe."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = openpyxl.load_workbook(embed_path)
    if SCOREBOARD_SHEET in wb.sheetnames:
        del wb[SCOREBOARD_SHEET]
    ws = wb.create_sheet(SCOREBOARD_SHEET)

    hdr_fill = PatternFill("solid", fgColor="0B6E4F")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin = Side(style="thin", color="D9D9D9")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)
    green = PatternFill("solid", fgColor="E6F4EA")
    red = PatternFill("solid", fgColor="FCE8E6")
    amber = PatternFill("solid", fgColor="FEF7E0")
    grey = PatternFill("solid", fgColor="F1F3F4")

    ws["A1"] = "Live Cohort Scoreboard — Phase 3-A out-of-sample guardrail"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (f"Auto-generated by guardrail_trace.py · source: Signals tab · as of "
                f"{as_of:%Y-%m-%d}. One row per live cohort. 'Held' = the guardrail's "
                "pairwise claim, 50-59% > 60-69% (distinct from the tracker's stricter "
                "'is 50-59 the top band' Pattern). Regime is keyed off the cohort-wide +5d "
                "hit rate: <40% compressed (bands can't separate), >55% clean — weight "
                "compressed cohorts lightly.")
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells("A2:I2")
    ws.row_dimensions[2].height = 54

    headers = ["Cohort", "n", "50-59% (h/n)", "50-59%", "Held (50-59>60-69)",
               "Cohort band order (by hit)", "Cum lead (pt)", "Cohort hit%", "Regime"]
    hr = 4
    for j, h in enumerate(headers, start=1):
        c = ws.cell(hr, j, h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        c.border = bd
    ws.row_dimensions[hr].height = 30

    if not rows:
        msg = ws.cell(hr + 1, 1, "No Phase 3-A cohorts have resolved their +5d yet.")
        msg.font = Font(italic=True, color="666666")
        ws.merge_cells(start_row=hr + 1, start_column=1, end_row=hr + 1, end_column=9)
    else:
        for i, r in enumerate(rows):
            rr = hr + 1 + i
            ws.cell(rr, 1, r["date"]).number_format = "m/d"
            ws.cell(rr, 2, r["n"]).alignment = Alignment(horizontal="center")
            ws.cell(rr, 3, (f"{r['l_h']} of {r['l_n']}" if r['l_n'] else "–")).alignment = \
                Alignment(horizontal="center")
            ws.cell(rr, 4, r["l_rate"]).number_format = "0.0%"
            hc = ws.cell(rr, 5, r["held"])
            hc.alignment = Alignment(horizontal="center")
            hc.font = Font(bold=True)
            hc.fill = green if r["held"] == "YES" else red if r["held"] == "NO" else grey
            ws.cell(rr, 6, r["order"]).alignment = Alignment(horizontal="left")
            gc = ws.cell(rr, 7, (round(r["cum_lead"] * 100, 1)
                                 if r["cum_lead"] is not None else None))
            gc.number_format = "0.0"
            ws.cell(rr, 8, r["cohort_hit"]).number_format = "0%"
            rc = ws.cell(rr, 9, r["regime"])
            rc.alignment = Alignment(horizontal="center")
            rc.fill = amber if r["regime"] == "compressed" else \
                green if r["regime"] == "clean" else grey
            for j in range(1, 10):
                ws.cell(rr, j).border = bd

        last = hr + len(rows)
        held_yes = sum(1 for r in rows if r["held"] == "YES")
        held_tot = sum(1 for r in rows if r["held"] in ("YES", "NO"))
        final_lead = next((r["cum_lead"] for r in reversed(rows)
                           if r["cum_lead"] is not None), None)
        sr = last + 2
        ws.cell(sr, 1, f"Pairwise held (50-59 > 60-69): {held_yes} of {held_tot} "
                       f"live cohorts").font = Font(bold=True)
        if final_lead is not None:
            ws.cell(sr + 1, 1, f"Cumulative lead through {as_of:%Y-%m-%d}: "
                               f"{final_lead * 100:+.1f}pt   ·   compressed cohorts are "
                               "low-information — weight them lightly")

    for j, w in enumerate([9, 5, 13, 8, 18, 28, 12, 11, 12], start=1):
        ws.column_dimensions[chr(64 + j)].width = w

    wb.save(embed_path)


def stratify_by_regime(cohorts, totals, phase3a_start):
    """Pool 50-59% vs 60-69% (+5d) outcomes WITHIN each regime stratum, split by
    phase (Phase 2 = before phase3a_start, Phase 3-A = on/after). Also pools a
    'non-compressed' stratum (mixed + clean) and an 'ALL' stratum per phase.

    Returns: {phase: {stratum: {LEAD:[h,n], RIVAL:[h,n], "_cohorts":[date,...]}}}
    The decisive cell for 're-litigating' Q1 is Phase-2 / non-compressed: if the
    50-59 lead is real there, the certified guardrail was not just a compressed
    artifact."""
    L, R = LEAD_BAND, RIVAL_BAND
    phases = ("Phase 2", "Phase 3-A")
    strat = {p: defaultdict(lambda: {L: [0, 0], R: [0, 0], "_cohorts": []})
             for p in phases}

    def add(phase, stratum, d, bands):
        cell = strat[phase][stratum]
        for b in (L, R):
            h, n = bands.get(b, [0, 0])
            cell[b][0] += h
            cell[b][1] += n
        cell["_cohorts"].append(d)

    for d, bands in cohorts:
        phase = "Phase 3-A" if d >= phase3a_start else "Phase 2"
        th, tn = totals.get(d, [0, 0])
        reg = regime_of((th / tn) if tn else None)
        add(phase, reg, d, bands)
        if reg in ("mixed", "clean"):
            add(phase, "non-compressed", d, bands)
        add(phase, "ALL", d, bands)
    return strat


def _stratum_stats(cell):
    """(lead_hits, lead_n, lead_rate, rival_hits, rival_n, rival_rate, lead_gap, n_cohorts)."""
    L, R = LEAD_BAND, RIVAL_BAND
    lh, ln = cell[L]
    rh, rn = cell[R]
    lr = (lh / ln) if ln else None
    rr = (rh / rn) if rn else None
    gap = (lr - rr) if (lr is not None and rr is not None) else None
    return lh, ln, lr, rh, rn, rr, gap, len(cell["_cohorts"])


# A non-compressed Phase-2 lead below this (in points) is treated as "not a real
# lead" -> the certified guardrail would then be a compressed-regime artifact.
ARTIFACT_THRESHOLD = 0.03


def verdict_lines(strat):
    """Plain-prose verdict sentences, shared by the console print and the .md."""
    def gap(phase, s):
        if s not in strat[phase]:
            return None
        return _stratum_stats(strat[phase][s])[6]
    p2_comp = gap("Phase 2", "compressed")
    p2_nc = gap("Phase 2", "non-compressed")
    p2_clean = gap("Phase 2", "clean")
    p2_mix = gap("Phase 2", "mixed")
    p2_all = gap("Phase 2", "ALL")
    p3_nc = gap("Phase 3-A", "non-compressed")
    p3_all = gap("Phase 3-A", "ALL")

    def pt(x):
        return "n/a" if x is None else f"{x * 100:+.1f}pt"

    out = []
    if p2_nc is None:
        out.append("Insufficient non-compressed Phase 2 data to rule on the question.")
        return out

    if p2_nc < ARTIFACT_THRESHOLD:
        out.append(f"Phase 2 50-59 lead in NON-COMPRESSED cohorts is only {pt(p2_nc)} "
                   f"(compressed {pt(p2_comp)}).")
        out.append("=> The certified +61.3% guardrail was largely a REGIME ARTIFACT, carried "
                   "by compressed/down cohorts. Its live failure is mechanically expected: the "
                   "50-59 > 60-69 ordering was never a calibration law in normal regimes, so it "
                   "does not generalize to the live (mostly non-compressed) cohorts.")
    else:
        out.append(f"Phase 2 50-59 STILL leads in NON-COMPRESSED cohorts: {pt(p2_nc)} "
                   f"(mixed {pt(p2_mix)}, clean {pt(p2_clean)}; compressed {pt(p2_comp)}, "
                   f"all-Phase-2 {pt(p2_all)}).")
        out.append("=> The guardrail was NOT merely a compressed-regime artifact: 50-59 led "
                   "60-69 across regimes in Phase 2, INCLUDING clean cohorts.")
        if p3_nc is not None:
            out.append(f"Yet within the SAME non-compressed regimes, Phase 3-A shows {pt(p3_nc)} "
                       f"(all-live {pt(p3_all)}).")
            if p3_nc < 0:
                out.append("The relationship has FLIPPED between Phase 2 and now within the same "
                           "regime stratum -> the more interesting story: SOMETHING CHANGED "
                           "between Phase 2 and live, and regime composition does not explain it. "
                           "Candidates: a market/volatility shift the hit-rate flag doesn't "
                           "capture, signal-generator drift between phases (checkable vs the "
                           "logs), temporal overfitting of Q1, or live small-sample noise.")
    out.append("NOTE: strata pool small n (especially live 60-69) — read leads as directional, "
               "weight by n, and let additional live cohorts firm the sign.")
    return out


_STRAT_ORDER = ["compressed", "mixed", "clean", "non-compressed", "ALL"]


def print_stratification(strat, phase3a_start):
    print("\n" + "=" * 72)
    print("PHASE 2 Q1 RE-EXAMINATION  —  50-59% vs 60-69% (+5d) WITHIN REGIME")
    print("=" * 72)
    print(f"Regime = cohort-wide +5d hit rate (compressed <{REGIME_COMPRESSED:.0%}, "
          f"clean >{REGIME_CLEAN:.0%}, else mixed). Split at {phase3a_start.isoformat()}.")
    for phase in ("Phase 2", "Phase 3-A"):
        print(f"\n  {phase}")
        print(f"    {'stratum':<15}{'coh':>4}  {'50-59% (n)':>14}  "
              f"{'60-69% (n)':>14}  {'lead':>9}")
        print("    " + "-" * 62)
        for s in _STRAT_ORDER:
            if s not in strat[phase]:
                continue
            lh, ln, lr, rh, rn, rr, gap, nc = _stratum_stats(strat[phase][s])
            ls = f"{lr:.1%} ({ln})" if lr is not None else f"-- ({ln})"
            rs = f"{rr:.1%} ({rn})" if rr is not None else f"-- ({rn})"
            gs = f"{gap * 100:+.1f}pt" if gap is not None else "--"
            mark = "  *" if s == "non-compressed" else ""
            print(f"    {s:<15}{nc:>4}  {ls:>14}  {rs:>14}  {gs:>9}{mark}")
    print("\n  " + "-" * 62)
    print("  VERDICT")
    for p in verdict_lines(strat):
        print("  " + p)


def write_stratification_md(strat, out_dir, as_of, phase3a_start):
    L, R = LEAD_BAND, RIVAL_BAND
    out = []
    out.append("# Phase 2 Q1 Re-Examination — 50-59% vs 60-69% by Regime")
    out.append("")
    out.append(f"*Generated {as_of} · regime = cohort-wide +5d hit rate "
               f"(compressed &lt;{REGIME_COMPRESSED:.0%}, clean &gt;{REGIME_CLEAN:.0%}, "
               f"else mixed) · phases split at {phase3a_start.isoformat()} · +5d "
               f"directional outcomes only*")
    out.append("")
    out.append("**Question.** Was the certified **+61.3%** Phase 2 lead of the 50-59% band a "
               "real, regime-robust calibration feature, or an artifact carried by "
               "compressed/down cohorts? The decisive cell is the **non-compressed** "
               "(mixed + clean) Phase 2 row.")
    out.append("")
    for phase in ("Phase 2", "Phase 3-A"):
        out.append(f"## {phase}")
        out.append("")
        out.append("| Stratum | Cohorts | 50-59% (n) | 60-69% (n) | Lead (50-59 − 60-69) |")
        out.append("|---|---:|---:|---:|---:|")
        for s in _STRAT_ORDER:
            if s not in strat[phase]:
                continue
            lh, ln, lr, rh, rn, rr, gap, nc = _stratum_stats(strat[phase][s])
            ls = f"{lr:.1%} ({ln})" if lr is not None else f"\u2013 ({ln})"
            rs = f"{rr:.1%} ({rn})" if rr is not None else f"\u2013 ({rn})"
            gs = f"**{gap * 100:+.1f}pt**" if gap is not None else "\u2013"
            label = f"**{s}**" if s == "non-compressed" else s
            out.append(f"| {label} | {nc} | {ls} | {rs} | {gs} |")
        out.append("")
    out.append("## Verdict")
    out.append("")
    for p in verdict_lines(strat):
        out.append(p.replace("=>", "**\u2192**"))
        out.append("")
    path = os.path.join(out_dir, "regime_stratification.md")
    with open(path, "w", newline="\n") as f:
        f.write("\n".join(out).rstrip() + "\n")
    return path


def main(argv=None):
    ap = argparse.ArgumentParser(description="Trace confidence-band guardrail "
                                             "compression cohort-by-cohort.")
    ap.add_argument("--tracker", default="claude_equity_bot_tracker.xlsx")
    ap.add_argument("--weekly", action="store_true",
                    help="aggregate cohorts into ISO weeks for the chart "
                         "(fixes the unbounded-x squish; per-cohort remains the default)")
    ap.add_argument("--since", default=None, help="YYYY-MM-DD inclusive")
    ap.add_argument("--until", default=None, help="YYYY-MM-DD inclusive")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--embed-into", default=None,
                    help="write/refresh a 'Guardrail Trace' sheet (data table + "
                         "chart image) into this .xlsx — pass a COPY of your tracker")
    ap.add_argument("--phase3a-start", default=PHASE3A_START.isoformat(),
                    help="YYYY-MM-DD first live cohort for the scoreboard "
                         f"(default {PHASE3A_START.isoformat()})")
    ap.add_argument("--stratify-regime", action="store_true",
                    help="also pool 50-59 vs 60-69 (+5d) WITHIN compressed/mixed/clean "
                         "regimes, split Phase 2 vs Phase 3-A; prints a table + verdict "
                         "and writes regime_stratification.md to --out-dir")
    args = ap.parse_args(argv)

    if not os.path.exists(args.tracker):
        sys.exit(f"ERROR: tracker not found: {args.tracker}")
    since = dt.date.fromisoformat(args.since) if args.since else None
    until = dt.date.fromisoformat(args.until) if args.until else None
    os.makedirs(args.out_dir, exist_ok=True)

    cohorts = load_cohorts(args.tracker, since, until)
    if not cohorts:
        sys.exit("ERROR: no resolved +5d signals in range.")
    table, series, dates = trace(cohorts)
    print_table(table)

    L, R = LEAD_BAND, RIVAL_BAND
    lc, rc, oc = series[L][-1], series[R][-1], series["40-49%"][-1]
    print("\n" + "=" * 60)
    print("FINAL CUMULATIVE (through %s):" % dates[-1])
    print(f"  50-59%: {lc:.1%}   60-69%: {rc:.1%}   40-49%: {oc:.1%}")
    print(f"  lead (50-59 over 60-69): {(lc - rc) * 100:+.1f}pt")
    print(f"  drift from certified 61.3%: {(lc - CERTIFIED_LEAD) * 100:+.1f}pt")
    # simple leveling check: compare gap over last 3 cohorts
    gaps = [series[L][i] - series[R][i] for i in range(len(dates))
            if series[L][i] is not None and series[R][i] is not None]
    if len(gaps) >= 4:
        recent = gaps[-3:]
        slope = recent[-1] - recent[0]
        trend = ("widening" if slope > 0.01 else
                 "narrowing" if slope < -0.01 else "roughly flat")
        print(f"  lead gap over last 3 cohorts: {trend} "
              f"({recent[0] * 100:+.1f}pt -> {recent[-1] * 100:+.1f}pt)")

    png = os.path.join(args.out_dir, "guardrail_trace.png")
    if args.weekly:
        wks = make_chart_weekly(cohorts, series, dates, png)
        print(f"  [weekly] {len(dates)} cohorts aggregated into {len(wks)} week(s)")
    else:
        make_chart(table, series, dates, png)
    print(f"\nWrote: {png}")

    # --- Live Cohort Scoreboard (Phase 3-A) ---
    phase3a_start = dt.date.fromisoformat(args.phase3a_start)
    totals = load_cohort_totals(args.tracker, since, until)
    sb_rows = build_scoreboard(cohorts, totals, series, dates, phase3a_start)
    print_scoreboard(sb_rows)

    if args.stratify_regime:
        strat = stratify_by_regime(cohorts, totals, phase3a_start)
        print_stratification(strat, phase3a_start)
        sp = write_stratification_md(strat, args.out_dir, dates[-1], phase3a_start)
        print(f"\nWrote: {sp}")

    if args.embed_into:
        if not os.path.exists(args.embed_into):
            sys.exit(f"ERROR: --embed-into target not found: {args.embed_into}")
        embed_into_workbook(table, series, dates, png, args.embed_into)
        print(f"Embedded '{SHEET_NAME}' sheet into: {args.embed_into}")
        embed_scoreboard_sheet(sb_rows, args.embed_into, dates[-1])
        print(f"Embedded '{SCOREBOARD_SHEET}' sheet into: {args.embed_into}")


if __name__ == "__main__":
    main()
