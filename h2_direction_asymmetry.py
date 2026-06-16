#!/usr/bin/env python3
"""
h2_direction_asymmetry.py  —  Claude Equity Bot, Phase 3-A · H2 success metric

Measures DIRECTION ASYMMETRY: does confidence move more per unit DOWN move
(de-rating into losses) than per unit UP move (re-rating into gains)? Phase 2
evidence was "supported (NVDA, WMT)" — NVDA was quick to de-rate on the drop and
never re-rated on the climb; GM updated 42->52 with much less responsiveness than
WMT — but the WMT +4.31% favorable-side breach complicates the pure-asymmetry
story. H2 adds symmetric-framing prompt language (variant B); the A/B should
pull the asymmetry ratio toward 1 WITHOUT breaking the +5d calibration.

THE METRIC (per ticker, then pooled)
    For each day-to-day step, take the signed price move and the |confidence move|.
        sens_down = sum |dConf| over DOWN steps / sum |dPrice%| over DOWN steps
        sens_up   = sum |dConf| over UP   steps / sum |dPrice%| over UP   steps
        asymmetry ratio = sens_down / sens_up   (pts of confidence per 1% move)
    ratio > 1  => more reactive to losses than gains (the documented bias)
    ratio ~ 1  => symmetric (the variant-B goal)

TWO MODES
    --tracker PATH (default)  baseline asymmetry from the Signals tab (Phase 2,
                              single arm). Establishes the ratio to beat.
    --logs GLOB               read variant-tagged 3-A logs, split by the JSON
                              prompt_variant field, and print the A-vs-B readout
                              (the actual A/B disentanglement).

USAGE
    python h2_direction_asymmetry.py
    python h2_direction_asymmetry.py --tracker claude_equity_bot_tracker.xlsx
    python h2_direction_asymmetry.py --logs "logs/signals_2026-06-*.log"
"""

from __future__ import annotations
import argparse
import datetime as dt
import glob
import json
import os
import re
import sys

SIGNALS_SHEET = "Signals"
FIRST_DATA_ROW = 5
COL = dict(date=1, ticker=2, signal=4, conf=5, p0=7, right5=14)
COHORT_START = dt.date(2026, 5, 18)
COHORT_END = dt.date(2026, 5, 29)
# ---- optional date-window slice (--since / --until) --------------------------
# When either bound is set, the analysis AND the calibration guardrail are
# computed over [SINCE, UNTIL]. With no bounds: analysis uses all rows and the
# guardrail uses the certified Phase 2 cohort window (COHORT_START..COHORT_END).
SINCE = None
UNTIL = None


def _in_window(d):
    if not isinstance(d, dt.date):
        return False
    if SINCE is not None and d < SINCE:
        return False
    if UNTIL is not None and d > UNTIL:
        return False
    return True


def _guardrail_window():
    if SINCE is not None or UNTIL is not None:
        return (SINCE or dt.date.min, UNTIL or dt.date.max)
    return (COHORT_START, COHORT_END)


def _window_label():
    if SINCE is not None or UNTIL is not None:
        return (str(SINCE) if SINCE else 'earliest') + ' .. ' + (str(UNTIL) if UNTIL else 'latest')
    return 'ALL ROWS (guardrail = Phase 2 cohort window)'

FOCUS = ("WMT", "GM", "NVDA")   # the names the Phase 2 asymmetry was argued on


def conf_pct(v):
    if v is None:
        return None
    return v * 100.0 if v <= 1.0 else float(v)


# --------------------------------------------------------------------------- #
# The asymmetry metric
# --------------------------------------------------------------------------- #
def asymmetry(seq):
    """seq: list of {conf, p0} in date order. Returns the directional split."""
    dd = du = cd = cu = 0.0
    dn = un = 0
    for i in range(1, len(seq)):
        a, b = seq[i - 1], seq[i]
        if a["conf"] is None or b["conf"] is None or not a["p0"] or not b["p0"]:
            continue
        dpx = (b["p0"] / a["p0"] - 1.0) * 100.0
        dconf = abs(b["conf"] - a["conf"])
        if dpx < 0:
            cd += dconf; dd += -dpx; dn += 1
        elif dpx > 0:
            cu += dconf; du += dpx; un += 1
    sens_down = (cd / dd) if dd else None
    sens_up = (cu / du) if du else None
    ratio = (sens_down / sens_up) if (sens_down and sens_up) else None
    return dict(down_n=dn, up_n=un, sens_down=sens_down, sens_up=sens_up, ratio=ratio)


def pool(seqs):
    """Pool many tickers into one aggregate asymmetry (sum, don't average ratios)."""
    dd = du = cd = cu = 0.0
    dn = un = 0
    for s in seqs:
        for i in range(1, len(s)):
            a, b = s[i - 1], s[i]
            if a["conf"] is None or b["conf"] is None or not a["p0"] or not b["p0"]:
                continue
            dpx = (b["p0"] / a["p0"] - 1.0) * 100.0
            dconf = abs(b["conf"] - a["conf"])
            if dpx < 0:
                cd += dconf; dd += -dpx; dn += 1
            elif dpx > 0:
                cu += dconf; du += dpx; un += 1
    sens_down = (cd / dd) if dd else None
    sens_up = (cu / du) if du else None
    ratio = (sens_down / sens_up) if (sens_down and sens_up) else None
    return dict(down_n=dn, up_n=un, sens_down=sens_down, sens_up=sens_up, ratio=ratio)


def f(x, d="—"):
    return d if x is None else f"{x:.2f}"


# --------------------------------------------------------------------------- #
# Loaders
# --------------------------------------------------------------------------- #
def load_tracker_by_ticker(tracker):
    import openpyxl
    wb = openpyxl.load_workbook(tracker, data_only=True)
    ws = wb[SIGNALS_SHEET]
    by = {}
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        d = ws.cell(row=r, column=COL["date"]).value
        d = d.date() if isinstance(d, dt.datetime) else d
        tk = ws.cell(row=r, column=COL["ticker"]).value
        if not _in_window(d) or not tk:
            continue
        by.setdefault(str(tk).strip().upper(), []).append(dict(
            date=d,
            conf=conf_pct(ws.cell(row=r, column=COL["conf"]).value),
            p0=ws.cell(row=r, column=COL["p0"]).value,
        ))
    wb.close()
    for tk in by:
        by[tk].sort(key=lambda x: x["date"])
    return by


def band_calibration(tracker):
    import openpyxl
    wb = openpyxl.load_workbook(tracker, data_only=True)
    ws = wb[SIGNALS_SHEET]
    agg = {"40-49": [0, 0], "50-59": [0, 0], "60-69": [0, 0]}
    glo, ghi = _guardrail_window()
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        d = ws.cell(row=r, column=COL["date"]).value
        d = d.date() if isinstance(d, dt.datetime) else d
        if not isinstance(d, dt.date) or not (glo <= d <= ghi):
            continue
        p = conf_pct(ws.cell(row=r, column=COL["conf"]).value)
        if p is None:
            continue
        b = ("40-49" if 40 <= p < 50 else "50-59" if 50 <= p < 60
             else "60-69" if 60 <= p < 70 else None)
        if b is None:
            continue
        agg[b][1] += 1
        if str(ws.cell(row=r, column=COL["right5"]).value or "").strip().upper() == "YES":
            agg[b][0] += 1
    wb.close()
    return agg


_JSON_RE = re.compile(r"Signal JSON:\s*\{")

def _extract_json(line):
    m = _JSON_RE.search(line)
    if not m:
        return None
    s = m.end() - 1
    depth = 0; instr = False; esc = False
    for i in range(s, len(line)):
        ch = line[i]
        if esc:
            esc = False; continue
        if ch == "\\":
            esc = True; continue
        if ch == '"':
            instr = not instr; continue
        if instr:
            continue
        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                return line[s:i + 1]
    return None


def load_logs_by_variant(globpat):
    """Returns {variant: {ticker: [ {date, conf, p0}, ... ]}}."""
    files = sorted(glob.glob(globpat))
    if not files:
        sys.exit(f"ERROR: no log files match {globpat!r}")
    out = {}
    for fp in files:
        date_str = os.path.basename(fp).replace("signals_", "").replace(".log", "")
        with open(fp, encoding="utf-8", errors="replace") as fh:
            for line in fh:
                raw = _extract_json(line)
                if not raw:
                    continue
                try:
                    p = json.loads(raw)
                except json.JSONDecodeError:
                    continue
                v = str(p.get("prompt_variant", "A")).upper()
                tk = str(p.get("ticker", "")).strip().upper()
                if not tk:
                    continue
                out.setdefault(v, {}).setdefault(tk, []).append(dict(
                    date=date_str, conf=conf_pct(p.get("confidence")),
                    p0=p.get("lastPrice"),
                ))
    for v in out:
        for tk in out[v]:
            out[v][tk].sort(key=lambda x: x["date"])
    return out, files


# --------------------------------------------------------------------------- #
# Reports
# --------------------------------------------------------------------------- #
def print_ratio_line(label, a):
    arrow = ""
    if a["ratio"] is not None:
        arrow = "  (de-rates into losses faster)" if a["ratio"] > 1.15 else \
                "  (re-rates into gains faster)" if a["ratio"] < 0.87 else \
                "  (~symmetric)"
    print(f"   {label:<22} sens_down={f(a['sens_down'])}  sens_up={f(a['sens_up'])}  "
          f"ratio={f(a['ratio'])}{arrow}")


def run_tracker(tracker, out_dir):
    by = load_tracker_by_ticker(tracker)
    if not by:
        sys.exit(f"ERROR: no signals in {tracker}")
    print("\n" + "=" * 82)
    print(f"H2 DIRECTION-ASYMMETRY — baseline (Signals tab: {os.path.basename(tracker)})")
    print("=" * 82)
    print("ratio = |dConf| per 1% DOWN move  /  |dConf| per 1% UP move   "
          "(>1 = more reactive to losses)\n")

    hdr = f"{'Ticker':<8}{'down':>5}{'up':>4}{'sens_down':>11}{'sens_up':>10}{'ratio':>8}"
    print(hdr); print("-" * 82)
    rows = []
    for tk in sorted(by):
        a = asymmetry(by[tk])
        rows.append((tk, a))
        print(f"{tk:<8}{a['down_n']:>5}{a['up_n']:>4}{f(a['sens_down']):>11}"
              f"{f(a['sens_up']):>10}{f(a['ratio']):>8}")

    print("\nFOCUS NAMES (the Phase 2 asymmetry was argued on these):")
    for tk in FOCUS:
        if tk in by:
            print_ratio_line(tk, asymmetry(by[tk]))

    agg = pool(list(by.values()))
    print("\nPORTFOLIO AGGREGATE (Phase 2 baseline):")
    print_ratio_line("portfolio", agg)

    bands = band_calibration(tracker)
    lead = max(bands, key=lambda b: (bands[b][0] / bands[b][1]) if bands[b][1] else -1)
    print("\nCALIBRATION GUARDRAIL — cohort-window +5d band hit rates (50-59% must lead):")
    for b in ("40-49", "50-59", "60-69"):
        h, n = bands[b]
        print(f"   {b}: {h}/{n} = {h/n*100:.1f}%" + ("  <-- leads" if b == lead else "")
              if n else f"   {b}: —")
    ok = lead == "50-59"
    print(f"   [{'PASS' if ok else 'WARN'}] 50-59% band {'leads' if ok else 'does NOT lead'}")

    print("\n" + "-" * 82)
    print("SUCCESS METRIC (the A/B, run with --logs once 3-A has A/B data)")
    print("-" * 82)
    rr = agg["ratio"]
    print(f"   • Variant B should pull the asymmetry ratio toward 1.00 "
          f"(baseline {f(rr)}).")
    print(f"   • sens_down and sens_up should converge (B re-rates into gains as")
    print(f"     readily as it de-rates into losses) — without flattening either.")
    print(f"   • The 50-59% +5d band keeps leading ({'holds' if ok else 'BROKEN'}).")
    print(f"   • Run the A/B both-arms day, then: "
          f"python {os.path.basename(__file__)} --logs \"logs/signals_*.log\"")

    md = [f"# H2 Direction-Asymmetry — baseline",
          f"*Source: `{os.path.basename(tracker)}` · generated {dt.datetime.now():%Y-%m-%d %H:%M}.*",
          f"*ratio = |dConf| per 1% down move / per 1% up move (>1 = more reactive to losses).*", "",
          f"**Portfolio ratio: {f(agg['ratio'])}** (sens_down {f(agg['sens_down'])} / sens_up {f(agg['sens_up'])})", "",
          "| Ticker | down | up | sens_down | sens_up | ratio |", "| --- | --- | --- | --- | --- | --- |"]
    for tk, a in rows:
        md.append(f"| {tk} | {a['down_n']} | {a['up_n']} | {f(a['sens_down'])} | "
                  f"{f(a['sens_up'])} | {f(a['ratio'])} |")
    path = os.path.join(out_dir, "h2_direction_asymmetry_baseline.md")
    open(path, "w", encoding="utf-8").write("\n".join(md) + "\n")
    print(f"\nWrote: {path}")


def run_logs(globpat, out_dir):
    data, files = load_logs_by_variant(globpat)
    print("\n" + "=" * 82)
    print(f"H2 DIRECTION-ASYMMETRY — A/B readout  ({len(files)} log file(s))")
    print("=" * 82)
    variants = sorted(data)
    if not variants:
        sys.exit("ERROR: no parseable signals in the logs.")
    results = {}
    for v in variants:
        agg = pool(list(data[v].values()))
        results[v] = agg
        n_tk = len(data[v])
        print(f"\nVARIANT {v}  ({n_tk} tickers):")
        print_ratio_line(f"variant {v}", agg)
        # focus names per variant
        for tk in FOCUS:
            if tk in data[v]:
                print_ratio_line(f"  {tk}", asymmetry(data[v][tk]))

    print("\n" + "-" * 82)
    print("A/B VERDICT")
    print("-" * 82)
    if "A" in results and "B" in results:
        ra, rb = results["A"]["ratio"], results["B"]["ratio"]
        if ra is not None and rb is not None:
            da, db = abs(ra - 1), abs(rb - 1)
            better = db < da
            print(f"   Variant A ratio = {f(ra)}  (|ratio-1| = {f(da)})")
            print(f"   Variant B ratio = {f(rb)}  (|ratio-1| = {f(db)})")
            print(f"   => {'B is MORE symmetric' if better else 'B is NOT more symmetric'} "
                  f"than A. {'Symmetric framing helps.' if better else 'Framing did not reduce the asymmetry.'}")
        else:
            print("   Not enough paired up/down moves yet to compute both ratios — "
                  "accrue more A/B sessions.")
    else:
        print(f"   Need both arms present; found variants: {', '.join(variants)}.")
    print("   (Pair with the tracker baseline + the 50-59% guardrail before acting.)")


def main(argv=None):
    ap = argparse.ArgumentParser(description="H2 direction-asymmetry / A/B success metric.")
    ap.add_argument("--tracker", default="claude_equity_bot_tracker.xlsx")
    ap.add_argument("--logs", default=None, help="glob of variant-tagged 3-A logs (A/B mode)")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--since", default=None,
                    help="only analyze rows on/after YYYY-MM-DD (e.g. 2026-06-15 = Phase 3-A only)")
    ap.add_argument("--until", default=None,
                    help="only analyze rows on/before YYYY-MM-DD")
    args = ap.parse_args(argv)
    global SINCE, UNTIL
    if args.since:
        SINCE = dt.date.fromisoformat(args.since)
    if args.until:
        UNTIL = dt.date.fromisoformat(args.until)
    if SINCE is not None or UNTIL is not None:
        print("")
        print("[DATE SLICE] analyzing " + _window_label())
        print("   (calibration guardrail uses this same window)")
    os.makedirs(args.out_dir, exist_ok=True)
    if args.logs:
        run_logs(args.logs, args.out_dir)
    else:
        if not os.path.exists(args.tracker):
            sys.exit(f"ERROR: tracker not found: {args.tracker}")
        run_tracker(args.tracker, args.out_dir)


if __name__ == "__main__":
    main()
