#!/usr/bin/env python3
# =============================================================================
# seed_signal_state.py
# =============================================================================
# PURPOSE:
#   One-shot (re-)build of signal_state.json from the tracker's Signals tab,
#   so the S1 prior-signal input is warm on day 1 post-deploy — and so the
#   state can be restored to the AM-canonical chain at any time (e.g. after a
#   PM diagnostic run overwrote a date's entry).
#
# USAGE:
#   python seed_signal_state.py
#   python seed_signal_state.py --tracker tracker_with_registry.xlsx
#   python seed_signal_state.py --tracker logs\claude_equity_bot_tracker.xlsx
#
# ANALYST NOTE:
#   Reads with load_workbook(data_only=True) (values, not formulas), scans
#   from row 5 with the 25-blank-row stop, and maps columns exactly as
#   h1_lag_trace.py does — one convention, two consumers. Confidence is
#   normalized to integer percent (the tracker stores fractions).
#
#   TWO-CHANNEL INTEGRITY (added after live Day 1, 2026-07-07): post-deploy
#   tracker rows carry the OPERATIVE (damped) confidence, but this state
#   file must carry RAW. The seed therefore overlays raw values recovered
#   from the logs' confidence_raw field (--logs glob, default
#   logs\signals_*.log): first occurrence per (date, ticker) wins
#   (AM-canonical), files whose names don't parse as signals_YYYY-MM-DD.log
#   (e.g. archived run-1 logs) are skipped by construction, and pre-deploy
#   logs contribute nothing (raw == operative there anyway). Net effect:
#   this one command is exactly-raw across eras and idempotent at any time.
# =============================================================================

import argparse
import datetime as dt
import glob
import os
import re

import openpyxl

from signal_state import save_signal_state, KEEP_LAST

SIGNALS_SHEET = "Signals"
FIRST_DATA_ROW = 5
COL = dict(date=1, ticker=2, signal=4, conf=5)


def as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    return v if isinstance(v, dt.date) else None


def conf_pct(v):
    if v is None:
        return None
    return v * 100.0 if v <= 1.0 else float(v)


def load_raw_overlay(pattern):
    """{(iso_date, TICKER): raw_conf_int} parsed from Signal JSON log lines.

    ANALYST NOTE:
        Mirrors h1_lag_trace.load_raw_conf_map: date comes from the
        filename; the FIRST occurrence per (date, ticker) wins, so a PM
        diagnostic appended to the same day's log can never shadow the AM
        raw value; unparseable filenames are silently skipped.
    """
    tick_re = re.compile(r'"ticker":\s*"([A-Za-z.\-]+)"')
    raw_re = re.compile(r'"confidence_raw":\s*(\d+)')
    m = {}
    for fp in sorted(glob.glob(pattern)):
        stem = os.path.splitext(os.path.basename(fp))[0]
        try:
            d = dt.date.fromisoformat(stem.replace("signals_", "")).isoformat()
        except ValueError:
            continue
        with open(fp, encoding="utf-8", errors="replace") as f:
            for line in f:
                if "Signal JSON" not in line:
                    continue
                tm, rm = tick_re.search(line), raw_re.search(line)
                if tm and rm:
                    key = (d, tm.group(1).upper())
                    if key not in m:
                        m[key] = int(rm.group(1))
    return m


def main():
    ap = argparse.ArgumentParser(
        description="Seed/restore signal_state.json from the tracker Signals tab.")
    ap.add_argument("--tracker", default="tracker_with_registry.xlsx")
    ap.add_argument("--out", default="signal_state.json")
    ap.add_argument("--last", type=int, default=KEEP_LAST,
                    help="sessions of history to keep per ticker (default 5)")
    ap.add_argument("--logs", default=os.path.join("logs", "signals_*.log"),
                    help="log glob for the RAW confidence overlay "
                         "(default: logs/signals_*.log)")
    args = ap.parse_args()

    overlay = load_raw_overlay(args.logs)
    n_overlaid = 0

    wb = openpyxl.load_workbook(args.tracker, data_only=True)
    ws = wb[SIGNALS_SHEET]

    per = {}
    blank = 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        d = as_date(ws.cell(row=r, column=COL["date"]).value)
        tk = ws.cell(row=r, column=COL["ticker"]).value
        if d is None and not tk:
            blank += 1
            if blank >= 25:
                break
            continue
        blank = 0
        if d is None or not tk:
            continue
        c = conf_pct(ws.cell(row=r, column=COL["conf"]).value)
        if c is None:
            continue
        sig = str(ws.cell(row=r, column=COL["signal"]).value or "").strip()
        tk_u = str(tk).strip().upper()
        raw = overlay.get((d.isoformat(), tk_u))
        if raw is not None:
            n_overlaid += 1
        per.setdefault(tk_u, []).append(
            {"date": d.isoformat(), "signal": sig,
             "conf": raw if raw is not None else int(round(c))})
    wb.close()

    state = {}
    for tk, entries in per.items():
        entries.sort(key=lambda e: e["date"])
        # ANALYST NOTE: collapse duplicate dates, last row of a date wins —
        # AM-canonical rows are one per ticker/day; duplicates would be
        # manual corrections, which sit lower in the sheet.
        dedup = {}
        for e in entries:
            dedup[e["date"]] = e
        state[tk] = [dedup[k] for k in sorted(dedup)][-args.last:]

    save_signal_state(state, args.out)
    print(f"Seeded {args.out}: {len(state)} ticker(s), "
          f"last {args.last} session(s) each.")
    print(f"  raw overlay from logs: {n_overlaid} value(s) applied "
          f"(0 is expected when no post-deploy rows are in seeding range)")
    for probe in ("WMT", "GM"):
        if probe in state:
            span = " | ".join(f"{e['date']}: {e['signal']} {e['conf']}%"
                              for e in state[probe])
            print(f"  {probe}: {span}")


if __name__ == "__main__":
    main()
