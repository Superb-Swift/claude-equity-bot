#!/usr/bin/env python3
"""
h4_dq_threshold.py  —  Claude Equity Bot, Phase 3-A · H4 watchdog

H4 (DQ-conditional KEEP threshold) is DEMOTED. The Phase 2 closeout: the
HIGH-DQ near-miss split finished 2/2 (SCHD, VTI) — real but n=2, far too small;
the same-ticker NVDA series went 2 wins / 2 losses (a coin flip); and broadly
"DQ measures input completeness, not directional accuracy" — HIGH-DQ HOLDs
(MSFT, PM) were among the biggest misses. So H4 is NOT a live change: there is
no threshold edit here and none is recommended. This script is the OBSERVATION
hook that says whether 3-A data ever reopens it.

It does four things, read-only on the tracker:
    1. DQ-stratified accuracy — does HIGH-DQ actually out-predict MEDIUM, across
       all resolved signals and within the 50-59% / 60-69% bands?
    2. The near-miss lever — BUY & 60-69% by DQ (where a DQ-conditional KEEP
       threshold would bite), with resolved hit rates.
    3. "DQ != thesis quality" check — the biggest HIGH-DQ +5d misses.
    4. A COUNTERFACTUAL — if HIGH-DQ near-miss BUYs had been kept at a lowered
       threshold, what would those extra kept signals have returned? (+ the n caveat.)
Then it applies a REOPENING RULE and prints the verdict, plus the 50-59% guardrail.

USAGE
    python h4_dq_threshold.py
    python h4_dq_threshold.py --tracker claude_equity_bot_tracker.xlsx
    python h4_dq_threshold.py --reopen-n 10 --reopen-hit 60
"""

from __future__ import annotations
import argparse
import datetime as dt
import os
import sys

import openpyxl

SIGNALS_SHEET = "Signals"
FIRST_DATA_ROW = 5
COL = dict(date=1, ticker=2, account=3, signal=4, conf=5, dq=6, p0=7,
           ret5=11, right5=14)
COHORT_START = dt.date(2026, 5, 18)
COHORT_END = dt.date(2026, 5, 29)
# ---- optional date-window slice (--since / --until) --------------------------
# When either bound is set, the analysis AND the calibration guardrail are
# computed over [SINCE, UNTIL]. With no bounds: analysis uses all rows and the
# guardrail uses the certified Phase 2 cohort window (COHORT_START..COHORT_END).
SINCE = None
UNTIL = None


def _in_window(d):
    if not isinstance(d, dt.date):
        return False
    if SINCE is not None and d < SINCE:
        return False
    if UNTIL is not None and d > UNTIL:
        return False
    return True


def _guardrail_window():
    if SINCE is not None or UNTIL is not None:
        return (SINCE or dt.date.min, UNTIL or dt.date.max)
    return (COHORT_START, COHORT_END)


def _window_label():
    if SINCE is not None or UNTIL is not None:
        return (str(SINCE) if SINCE else 'earliest') + ' .. ' + (str(UNTIL) if UNTIL else 'latest')
    return 'ALL ROWS (guardrail = Phase 2 cohort window)'

DQ_TIERS = ("HIGH", "MEDIUM", "LOW")


def as_date(v):
    if isinstance(v, dt.datetime):
        return v.date()
    return v if isinstance(v, dt.date) else None


def conf_pct(v):
    if v is None:
        return None
    return v * 100.0 if v <= 1.0 else float(v)


def num(v):
    return float(v) if isinstance(v, (int, float)) else None


def pct_num(v):
    """Return+Nd columns are stored as fractions (-0.116 = -11.6%); to percent."""
    return float(v) * 100.0 if isinstance(v, (int, float)) else None


def fmt(x, nd=1, suffix="", dash="—"):
    return dash if x is None else f"{x:.{nd}f}{suffix}"


def load_rows(tracker):
    wb = openpyxl.load_workbook(tracker, data_only=True)
    ws = wb[SIGNALS_SHEET]
    rows = []
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        d = as_date(ws.cell(row=r, column=COL["date"]).value)
        tk = ws.cell(row=r, column=COL["ticker"]).value
        if not _in_window(d) or not tk:
            continue
        rows.append(dict(
            date=d, ticker=str(tk).strip().upper(),
            signal=str(ws.cell(row=r, column=COL["signal"]).value or "").strip().upper(),
            conf=conf_pct(ws.cell(row=r, column=COL["conf"]).value),
            dq=str(ws.cell(row=r, column=COL["dq"]).value or "").strip().upper(),
            ret5=pct_num(ws.cell(row=r, column=COL["ret5"]).value),
            right5=str(ws.cell(row=r, column=COL["right5"]).value or "").strip().upper(),
        ))
    wb.close()
    return rows


def hit_stats(subset):
    """(hits, resolved_n, hit_rate, avg_ret5 over resolved) for rows with YES/NO."""
    res = [x for x in subset if x["right5"] in ("YES", "NO")]
    hits = sum(1 for x in res if x["right5"] == "YES")
    rets = [x["ret5"] for x in res if x["ret5"] is not None]
    rate = (hits / len(res) * 100.0) if res else None
    avg = (sum(rets) / len(rets)) if rets else None
    return hits, len(res), rate, avg


def run(tracker, reopen_n, reopen_hit, out_dir):
    rows = load_rows(tracker)
    if not rows:
        sys.exit(f"ERROR: no signals in {tracker}")

    print("\n" + "=" * 80)
    print(f"H4 WATCHDOG — DQ-conditional threshold (DEMOTED)   "
          f"source: {os.path.basename(tracker)}")
    print("=" * 80)
    print("H4 is on the backlog. No threshold change is made or recommended here;")
    print("this is the monitor that reports whether the data reopens it.\n")

    # ---- 1. DQ distribution ----
    from collections import Counter
    dist = Counter(x["dq"] or "(blank)" for x in rows)
    total = len(rows)
    print(f"DQ DISTRIBUTION ({total} signal rows):")
    for tier in DQ_TIERS:
        c = dist.get(tier, 0)
        print(f"   {tier:<7}: {c:>4}  ({c/total*100:.1f}%)")
    other = total - sum(dist.get(t, 0) for t in DQ_TIERS)
    if other:
        print(f"   other  : {other:>4}  ({other/total*100:.1f}%)")

    # ---- 2. broad DQ-accuracy test ----
    print("\nDQ-STRATIFIED +5d ACCURACY (all resolved signals; is HIGH better?):")
    print(f"   {'DQ':<7}{'hits/res':>10}{'hit%':>8}{'avg +5d ret':>13}")
    dq_overall = {}
    for tier in DQ_TIERS:
        h, n, rate, avg = hit_stats([x for x in rows if x["dq"] == tier])
        dq_overall[tier] = (rate, n)
        print(f"   {tier:<7}{f'{h}/{n}':>10}{fmt(rate,1,'%'):>8}{fmt(avg,2,'%'):>13}")

    print("\n   within bands, by DQ (the calibration premise of a DQ-conditional rule):")
    for lo, hi, name in ((50, 60, "50-59"), (60, 70, "60-69")):
        print(f"     band {name}:")
        for tier in DQ_TIERS:
            sub = [x for x in rows if x["dq"] == tier and x["conf"] is not None
                   and lo <= x["conf"] < hi]
            h, n, rate, avg = hit_stats(sub)
            if n:
                print(f"       {tier:<7}{f'{h}/{n}':>8}{fmt(rate,1,'%'):>8}")

    # ---- 3. near-miss lever ----
    print("\nNEAR-MISS LEVER — BUY & 60-69% by DQ (where a DQ-conditional KEEP bites):")
    nm = [x for x in rows if x["signal"] == "BUY" and x["conf"] is not None
          and 60 <= x["conf"] < 70]
    print(f"   {'DQ':<7}{'total':>7}{'hits/res':>10}{'hit%':>8}{'avg +5d ret':>13}")
    nm_high = None
    for tier in DQ_TIERS:
        sub = [x for x in nm if x["dq"] == tier]
        h, n, rate, avg = hit_stats(sub)
        if tier == "HIGH":
            nm_high = (h, n, rate, avg)
        if sub:
            print(f"   {tier:<7}{len(sub):>7}{f'{h}/{n}':>10}{fmt(rate,1,'%'):>8}{fmt(avg,2,'%'):>13}")
    print(f"   ({len(nm)} near-miss BUYs total — cross-checks the Q2 registry of 17.)")

    # ---- 4. DQ != thesis quality ----
    print("\nDQ != THESIS QUALITY — biggest HIGH-DQ +5d misses (input completeness")
    print("is not directional accuracy; cf. MSFT / PM high-DQ HOLD misses):")
    high_miss = sorted([x for x in rows if x["dq"] == "HIGH" and x["ret5"] is not None],
                       key=lambda x: x["ret5"])[:5]
    for x in high_miss:
        print(f"   {x['date']}  {x['ticker']:<5} {x['signal']:<4} "
              f"conf {fmt(x['conf'],0)}%  +5d {fmt(x['ret5'],2,'%')}")

    # ---- 5. counterfactual ----
    print("\nCOUNTERFACTUAL — keep HIGH-DQ near-miss BUYs at a lowered threshold:")
    if nm_high and nm_high[1]:
        h, n, rate, avg = nm_high
        print(f"   The {len([x for x in nm if x['dq']=='HIGH'])} HIGH-DQ near-miss BUYs "
              f"would become KEEPs; resolved {h}/{n} = {fmt(rate,1,'%')}, "
              f"avg +5d {fmt(avg,2,'%')}.")
        print(f"   Looks favorable, but n={n}. Not actionable on this sample.")
    else:
        print("   No resolved HIGH-DQ near-miss BUYs to simulate.")

    # ---- reopening rule ----
    print("\n" + "-" * 80)
    print("REOPENING RULE — H4 leaves the backlog ONLY if ALL hold:")
    print("-" * 80)
    h_rate, h_n = (nm_high[2], nm_high[1]) if nm_high else (None, 0)
    hi_ov = dq_overall.get("HIGH", (None, 0))[0]
    med_ov = dq_overall.get("MEDIUM", (None, 0))[0]
    c1 = h_n >= reopen_n
    c2 = (h_rate is not None) and (h_rate >= reopen_hit)
    c3 = (hi_ov is not None and med_ov is not None) and (hi_ov >= med_ov)
    print(f"   [{'PASS' if c1 else 'FAIL'}] HIGH-DQ near-miss resolved n ≥ {reopen_n}  "
          f"(now {h_n})")
    print(f"   [{'PASS' if c2 else 'FAIL'}] HIGH-DQ near-miss hit rate ≥ {reopen_hit}%  "
          f"(now {fmt(h_rate,1,'%')})")
    print(f"   [{'PASS' if c3 else 'FAIL'}] HIGH-DQ overall accuracy ≥ MEDIUM  "
          f"(HIGH {fmt(hi_ov,1,'%')} vs MEDIUM {fmt(med_ov,1,'%')})")
    reopen = c1 and c2 and c3
    print(f"\n   VERDICT: H4 {'REOPENS — re-scope a DQ-conditional threshold' if reopen else 'STAYS DEMOTED (backlog)'}.")
    if not reopen:
        print("   The binding gap is sample size on HIGH-DQ near-misses; only a larger")
        print("   3-A near-miss sample can reopen it. Re-run as 3-A near-misses accrue.")

    # ---- guardrail ----
    agg = {"40-49": [0, 0], "50-59": [0, 0], "60-69": [0, 0]}
    glo, ghi = _guardrail_window()
    for x in rows:
        if not (glo <= x["date"] <= ghi) or x["conf"] is None:
            continue
        b = ("40-49" if 40 <= x["conf"] < 50 else "50-59" if 50 <= x["conf"] < 60
             else "60-69" if 60 <= x["conf"] < 70 else None)
        if b is None:
            continue
        agg[b][1] += 1
        if x["right5"] == "YES":
            agg[b][0] += 1
    lead = max(agg, key=lambda b: (agg[b][0] / agg[b][1]) if agg[b][1] else -1)
    print("\nCALIBRATION GUARDRAIL — cohort-window +5d bands (50-59% must lead):")
    for b in ("40-49", "50-59", "60-69"):
        h, n = agg[b]
        print(f"   {b}: {h}/{n} = {h/n*100:.1f}%" + ("  <-- leads" if b == lead else "")
              if n else f"   {b}: —")
    ok = lead == "50-59"
    print(f"   [{'PASS' if ok else 'WARN'}] 50-59% band {'leads' if ok else 'does NOT lead'}")

    # ---- markdown ----
    md = [f"# H4 Watchdog — DQ-conditional threshold (DEMOTED)",
          f"*Source: `{os.path.basename(tracker)}` · generated {dt.datetime.now():%Y-%m-%d %H:%M}.*",
          "", "**H4 is on the backlog — no threshold change is made or recommended.** "
          "This is the monitor for whether 3-A data reopens it.", "",
          "## DQ-stratified +5d accuracy (all resolved)",
          "| DQ | hits/res | hit% | avg +5d ret |", "| --- | --- | --- | --- |"]
    for tier in DQ_TIERS:
        h, n, rate, avg = hit_stats([x for x in rows if x["dq"] == tier])
        md.append(f"| {tier} | {h}/{n} | {fmt(rate,1,'%')} | {fmt(avg,2,'%')} |")
    md += ["", "## Near-miss lever — BUY & 60-69% by DQ",
           "| DQ | total | hits/res | hit% |", "| --- | --- | --- | --- |"]
    for tier in DQ_TIERS:
        sub = [x for x in nm if x["dq"] == tier]
        h, n, rate, avg = hit_stats(sub)
        if sub:
            md.append(f"| {tier} | {len(sub)} | {h}/{n} | {fmt(rate,1,'%')} |")
    md += ["", f"## Verdict", f"H4 {'REOPENS' if reopen else 'STAYS DEMOTED (backlog)'} "
           f"— reopening needs HIGH-DQ near-miss n ≥ {reopen_n} (now {h_n}), "
           f"hit ≥ {reopen_hit}% (now {fmt(h_rate,1,'%')}), and HIGH ≥ MEDIUM overall."]
    path = os.path.join(out_dir, "h4_dq_threshold_baseline.md")
    open(path, "w", encoding="utf-8").write("\n".join(md) + "\n")
    print(f"\nWrote: {path}")


def main(argv=None):
    ap = argparse.ArgumentParser(description="H4 DQ-conditional-threshold watchdog.")
    ap.add_argument("--tracker", default="claude_equity_bot_tracker.xlsx")
    ap.add_argument("--reopen-n", type=int, default=10,
                    help="min resolved HIGH-DQ near-misses to consider reopening (default 10)")
    ap.add_argument("--reopen-hit", type=float, default=60.0,
                    help="min HIGH-DQ near-miss hit%% to consider reopening (default 60)")
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--since", default=None,
                    help="only analyze rows on/after YYYY-MM-DD (e.g. 2026-06-15 = Phase 3-A only)")
    ap.add_argument("--until", default=None,
                    help="only analyze rows on/before YYYY-MM-DD")
    args = ap.parse_args(argv)
    global SINCE, UNTIL
    if args.since:
        SINCE = dt.date.fromisoformat(args.since)
    if args.until:
        UNTIL = dt.date.fromisoformat(args.until)
    if SINCE is not None or UNTIL is not None:
        print("")
        print("[DATE SLICE] analyzing " + _window_label())
        print("   (calibration guardrail uses this same window)")
    if not os.path.exists(args.tracker):
        sys.exit(f"ERROR: tracker not found: {args.tracker}")
    os.makedirs(args.out_dir, exist_ok=True)
    run(args.tracker, args.reopen_n, args.reopen_hit, args.out_dir)


if __name__ == "__main__":
    main()
