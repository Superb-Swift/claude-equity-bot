#!/usr/bin/env python3
# =============================================================================
# monitor_strip.py  —  the OPEN QUESTIONS dashboard (daily glance)
# =============================================================================
# WHY THIS EXISTS
#   guardrail_trace.py answers "does 50-59 lead 60-69?" — a question CLOSED by
#   the Q1-R diagnostic (guardrail retired, T2 retired). This strip tracks the
#   questions that are still OPEN, in fixed width regardless of how much
#   history accumulates:
#
#     WS1 / G2   blend vs base        gate-flip net + confirm/discount arms
#     WS2 / G1   H1 acceptance clock  episodes detected + response verdict
#     H4         DQ-conditional gate  HIGH-DQ near-miss n & hit% vs the bars
#     KEEP-70    the 70% BUY gate     all resolved near-miss BUYs
#     context    registry size, live band split
#
# EPISODE DETECTION (WS2) — the one encoded judgement, and it is NOT a new
#   number: "material adverse move" reuses D1's pre-registered
#   DampingConfig.THETA (5.0%), imported from risk_engine so there is exactly
#   ONE definition of "material" in the system. An episode = a tracer session
#   whose trailing 5-day move is <= -THETA. It is SERVED if the operative
#   confidence falls within RESPONSE_SESSIONS (2) sessions of that move; the
#   acceptance criterion (WS2_H1_Feature_Spec / G1) is the table read, and this
#   counter makes "how many test cases have we actually had" reproducible
#   rather than a recollection.
#
# DATA SOURCES
#   tracker Signals tab  — bands, near-miss/H4, KEEP-70, tracer confidences
#   logs/signals_*.log   — confidence_blend (the blend channel lives only in
#                          the logs; there is no tracker column by design)
#   Returns are computed from the Price columns, so an unrecalculated formula
#   cache cannot affect this tool.
#
# USAGE
#   python monitor_strip.py --tracker tracker_with_registry.xlsx
#   python monitor_strip.py --tracker ... --logs "logs\\signals_*.log" --no-png
# =============================================================================

from __future__ import annotations
import argparse
import datetime as dt
import glob
import json
import os
import re
from collections import defaultdict

import openpyxl

SIGNALS_SHEET = "Signals"
FIRST_DATA_ROW = 5
COL = dict(date=1, ticker=2, account=3, signal=4, conf=5, dq=6, p0=7, p5=8)

LIVE_START = dt.date(2026, 6, 15)     # Phase 3-A go-live (population boundary)
WS2_DEPLOY = dt.date(2026, 7, 7)      # S1+D1 go-live (era A-S1D1) — G1 clock start
WS1_DEPLOY = dt.date(2026, 7, 13)     # blend go-live (era A-S1D1-B1)
TRACERS = ("WMT", "GM")               # WS2 primary / secondary
RESPONSE_SESSIONS = 2                 # G1: confidence must respond within 2
HOLD_BAND = 0.03                      # +/-3% -> a HOLD was "right"
BUY_GATE = 70                         # KEEP-70
H4_MIN_N = 10                         # H4 reopen: resolved HIGH-DQ near-misses
H4_MIN_HIT = 60.0                     # H4 reopen: hit rate bar

try:                                   # one definition of "material adverse"
    from risk_engine import DampingConfig
    THETA = float(DampingConfig.THETA)
    THETA_SRC = "risk_engine.DampingConfig.THETA"
except Exception:                      # standalone fallback, kept in sync
    THETA = 5.0
    THETA_SRC = "local default (risk_engine unavailable)"


# ----------------------------------------------------------------- loading --
def _date(v):
    return v.date() if isinstance(v, dt.datetime) else (v if isinstance(v, dt.date) else None)


def _num(v):
    return v if isinstance(v, (int, float)) else None


def load_rows(tracker):
    wb = openpyxl.load_workbook(tracker, data_only=True)
    ws = wb[SIGNALS_SHEET]
    rows, blank = [], 0
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        d = _date(ws.cell(r, COL["date"]).value)
        tk = ws.cell(r, COL["ticker"]).value
        if d is None and not tk:
            blank += 1
            if blank >= 30:
                break
            continue
        blank = 0
        if d is None or not tk:
            continue
        conf = ws.cell(r, COL["conf"]).value
        if isinstance(conf, (int, float)) and conf <= 1:
            conf *= 100
        p0, p5 = _num(ws.cell(r, COL["p0"]).value), _num(ws.cell(r, COL["p5"]).value)
        rows.append(dict(
            date=d, ticker=str(tk).strip().upper(),
            signal=str(ws.cell(r, COL["signal"]).value or "").strip().upper(),
            conf=conf, dq=str(ws.cell(r, COL["dq"]).value or "").strip().upper(),
            p0=p0, ret5=((p5 - p0) / p0 if (p0 and p5) else None),
        ))
    wb.close()
    rows.sort(key=lambda x: (x["date"], x["ticker"]))
    return rows


def load_blend(pattern):
    """{(date,ticker): {base, blend}} from the logs (blend lives only there)."""
    out = {}
    for fp in sorted(glob.glob(pattern)):
        m = re.search(r"signals_(\d{4}-\d{2}-\d{2})", os.path.basename(fp))
        if not m:
            continue                                  # archived/renamed logs
        d = dt.date.fromisoformat(m.group(1))
        for line in open(fp, encoding="utf-8", errors="replace"):
            if "Signal JSON" not in line:
                continue
            j = line.split("Signal JSON:", 1)[1]
            j = j[:j.rfind("} |") + 1] if "} |" in j else j
            try:
                p = json.loads(j)
            except Exception:
                continue
            if "confidence_blend" not in p:
                continue
            key = (d, str(p.get("ticker", "")).upper())
            out.setdefault(key, dict(base=p.get("confidence"),
                                     blend=p.get("confidence_blend"),
                                     signal=str(p.get("signal", "")).upper()))
    return out


# ----------------------------------------------------------------- metrics --
def right5(sig, ret):
    if ret is None:
        return None
    return abs(ret) < HOLD_BAND if sig == "HOLD" else (ret > 0 if sig == "BUY" else ret < 0)


def _acc(sub):
    sub = [x for x in sub if x["ret5"] is not None]
    h = sum(1 for x in sub if right5(x["signal"], x["ret5"]))
    return len(sub), (100.0 * h / len(sub) if sub else 0.0)


def ws2_episodes(rows):
    """WS2/G1 — count post-deploy tracer episodes and whether each was SERVED.

    episode  : trailing 5-session move <= -THETA on a tracer
    served   : operative confidence FALLS within RESPONSE_SESSIONS sessions
    """
    out = {}
    for tk in TRACERS:
        ser = [x for x in rows if x["ticker"] == tk and x["p0"]]
        ser.sort(key=lambda x: x["date"])
        post = [i for i, x in enumerate(ser) if x["date"] >= WS1_DEPLOY or True]  # full series indexable
        pre_eps = post_eps = post_served = 0
        last_ep = None
        for i, x in enumerate(ser):
            if i < 5 or x["date"] < LIVE_START:
                continue
            trail = (x["p0"] - ser[i - 5]["p0"]) / ser[i - 5]["p0"] * 100.0
            if trail > -THETA:
                continue
            last_ep = x["date"]
            if x["date"] < WS2_DEPLOY:          # baseline era — not the G1 test
                pre_eps += 1
                continue
            post_eps += 1                        # THIS is the acceptance clock
            base = x["conf"]
            window = ser[i + 1: i + 1 + RESPONSE_SESSIONS]
            if base is not None and any(
                    (w["conf"] is not None and w["conf"] < base) for w in window):
                post_served += 1
        post_sessions = sum(1 for x in ser if x["date"] >= WS2_DEPLOY)
        sessions_since = (sum(1 for x in ser if x["date"] > last_ep)
                          if last_ep is not None else None)
        out[tk] = dict(episodes=post_eps, served=post_served, pre=pre_eps,
                       last=last_ep, since=sessions_since,
                       post_sessions=post_sessions, n=len(ser))
    return out


def ws1_blend(rows, blend):
    """WS1/G2 — gate-flip net and confirm/discount arms (post-deploy only)."""
    idx = {(x["date"], x["ticker"]): x for x in rows}
    confirm, discount, flips, net = [], [], [], 0
    adjusted = 0
    for (d, tk), b in blend.items():
        if d < WS1_DEPLOY:
            continue
        base, bl = b.get("base"), b.get("blend")
        if not isinstance(base, (int, float)) or not isinstance(bl, (int, float)):
            continue
        if bl != base:
            adjusted += 1
        row = idx.get((d, tk))
        if row is None or row["ret5"] is None:
            continue
        if b["signal"] in ("BUY", "SELL"):
            if bl > base:
                confirm.append(row)
            elif bl < base:
                discount.append(row)
        if b["signal"] == "BUY" and ((base >= BUY_GATE) != (bl >= BUY_GATE)):
            good = (row["ret5"] > 0) == (bl >= BUY_GATE)
            net += 1 if good else -1
            flips.append((d, tk, base, bl, row["ret5"], good))
    nc, ac = _acc(confirm)
    nd, ad = _acc(discount)
    return dict(adjusted=adjusted, n_confirm=nc, acc_confirm=ac,
                n_discount=nd, acc_discount=ad, flips=flips, net=net)


def near_miss(rows):
    nm = [x for x in rows if x["signal"] == "BUY"
          and isinstance(x["conf"], (int, float)) and 60 <= x["conf"] < BUY_GATE]
    res = [x for x in nm if x["ret5"] is not None]
    hits = sum(1 for x in res if right5(x["signal"], x["ret5"]))
    hi = [x for x in nm if x["dq"] == "HIGH"]
    hres = [x for x in hi if x["ret5"] is not None]
    hhits = sum(1 for x in hres if right5(x["signal"], x["ret5"]))
    return dict(total=len(nm), resolved=len(res), hits=hits,
                rate=(100.0 * hits / len(res) if res else 0.0),
                hi_total=len(hi), hi_resolved=len(hres), hi_hits=hhits,
                hi_rate=(100.0 * hhits / len(hres) if hres else 0.0))


def collect(tracker, logs):
    rows = load_rows(tracker)
    blend = load_blend(logs) if logs else {}
    live = [x for x in rows if x["date"] >= LIVE_START]
    n59, a59 = _acc([x for x in live if isinstance(x["conf"], (int, float)) and 50 <= x["conf"] < 60])
    n69, a69 = _acc([x for x in live if isinstance(x["conf"], (int, float)) and 60 <= x["conf"] < 70])
    nH, aH = _acc([x for x in live if x["dq"] == "HIGH"])
    nM, aM = _acc([x for x in live if x["dq"] == "MEDIUM"])
    return dict(
        as_of=max(x["date"] for x in rows), n_rows=len(rows),
        ws1=ws1_blend(rows, blend), ws2=ws2_episodes(rows), nm=near_miss(rows),
        bands=dict(n59=n59, a59=a59, n69=n69, a69=a69),
        dq=dict(nH=nH, aH=aH, nM=nM, aM=aM),
    )


# ------------------------------------------------------------------ output --
def to_markdown(m):
    ws1, ws2, nm, b, dq = m["ws1"], m["ws2"], m["nm"], m["bands"], m["dq"]
    h4_n_ok = nm["hi_resolved"] >= H4_MIN_N
    h4_r_ok = nm["hi_rate"] >= H4_MIN_HIT
    L = []
    L.append("# Monitor strip — open questions")
    L.append(f"\n*Source: tracker Signals tab + logs · as of {m['as_of']} · "
             f"{m['n_rows']} signal rows. Returns computed from the Price columns "
             f"(cache-independent). Episode threshold THETA={THETA:.1f}% from {THETA_SRC}.*\n")
    L.append("| Question | Reading | Bar | State |")
    L.append("|---|---|---|---|")
    L.append(f"| **WS1 / G2** blend vs base | gate-flip net **{ws1['net']:+d}** "
             f"({len(ws1['flips'])} flip(s)); confirm {ws1['acc_confirm']:.0f}% (n={ws1['n_confirm']}) "
             f"vs discount {ws1['acc_discount']:.0f}% (n={ws1['n_discount']}) "
             f"| net ≥ 0 & confirm > discount, ≥5 cohorts | "
             f"{'on track' if ws1['net'] >= 0 else 'negative'} |")
    for tk in TRACERS:
        e = ws2[tk]
        last = e["last"].isoformat() if e["last"] else "none"
        detail = (f"**{e['episodes']} post-deploy episode(s)** in {e['post_sessions']} session(s)"
                  f" ({e['served']} served ≤{RESPONSE_SESSIONS}); "
                  f"{e['pre']} pre-deploy; last episode {last}"
                  + (f", {e['since']} session(s) ago" if e['since'] is not None else ""))
        L.append(f"| **WS2 / G1** {tk} clock | {detail} "
                 f"| ≥1 episode with response ≤{RESPONSE_SESSIONS} sessions | "
                 f"{'readable' if e['episodes'] else 'NO TEST CASES YET'} |")
    L.append(f"| **H4** DQ-conditional gate | HIGH-DQ near-miss **{nm['hi_rate']:.0f}%** "
             f"({nm['hi_hits']}/{nm['hi_resolved']} resolved, {nm['hi_total']} total) "
             f"| n ≥ {H4_MIN_N} resolved AND hit ≥ {H4_MIN_HIT:.0f}% | "
             f"{'REOPEN' if (h4_n_ok and h4_r_ok) else 'demoted'} "
             f"({'n ok' if h4_n_ok else 'n short'}, {'rate ok' if h4_r_ok else 'rate FAILING'}) |")
    L.append(f"| **KEEP-70** the BUY gate | rejected BUYs that would have paid: "
             f"**{nm['rate']:.1f}%** ({nm['hits']}/{nm['resolved']}) "
             f"| < 50% supports keeping 70 | "
             f"{'supported' if nm['rate'] < 50 else 'review'} |")
    L.append("")
    L.append("**Context** — "
             f"registry {nm['total']} near-miss BUYs · "
             f"live bands 50-59 {b['a59']:.1f}% (n={b['n59']}) vs 60-69 {b['a69']:.1f}% (n={b['n69']}) · "
             f"DQ HIGH {dq['aH']:.1f}% (n={dq['nH']}) vs MEDIUM {dq['aM']:.1f}% (n={dq['nM']}).")
    if ws1["flips"]:
        L.append("\n**Gate-flips (WS1):**\n")
        L.append("| Date | Ticker | Base | Blend | +5d | Blend better? |")
        L.append("|---|---|---:|---:|---:|---|")
        for d, tk, base, bl, ret, good in ws1["flips"]:
            L.append(f"| {d} | {tk} | {base} | {bl} | {ret*100:+.2f}% | "
                     f"{'yes' if good else 'no'} |")
    L.append("\n*Advisory-only · no orders · not financial advice.*")
    return "\n".join(L)


def make_png(m, out_path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt

    ws1, ws2, nm, b = m["ws1"], m["ws2"], m["nm"], m["bands"]
    h4_ok = (nm["hi_resolved"] >= H4_MIN_N) and (nm["hi_rate"] >= H4_MIN_HIT)
    eps = sum(ws2[t]["episodes"] for t in TRACERS)
    cards = [
        ("WS1 blend (G2)", f"{ws1['net']:+d}", "gate-flip net",
         "#1a73e8" if ws1["net"] >= 0 else "#d93025"),
        ("WS2 clock (G1)", f"{eps}", f"episodes (theta {THETA:.0f}%)",
         "#1a73e8" if eps else "#d93025"),
        ("H4 near-miss", f"{nm['hi_rate']:.0f}%", f"{nm['hi_resolved']}/{H4_MIN_N} · bar {H4_MIN_HIT:.0f}%",
         "#1a73e8" if h4_ok else "#9aa0a6"),
        ("KEEP-70", f"{nm['rate']:.0f}%", f"{nm['hits']}/{nm['resolved']} paid",
         "#1a73e8" if nm["rate"] < 50 else "#d93025"),
    ]
    fig, axes = plt.subplots(1, len(cards), figsize=(12, 2.6))
    for ax, (label, value, sub, color) in zip(axes, cards):
        ax.axis("off")
        ax.add_patch(plt.Rectangle((0, 0), 1, 1, transform=ax.transAxes,
                                   facecolor="#f7f7f5", edgecolor="none"))
        ax.text(0.5, 0.78, label, ha="center", va="center", fontsize=10,
                color="#5f5e5a", transform=ax.transAxes)
        ax.text(0.5, 0.46, value, ha="center", va="center", fontsize=26,
                color=color, transform=ax.transAxes)
        ax.text(0.5, 0.16, sub, ha="center", va="center", fontsize=8.5,
                color="#898781", transform=ax.transAxes)
    fig.suptitle(f"Open questions — as of {m['as_of']}   |   "
                 f"registry {nm['total']} · live 50-59 {b['a59']:.0f}% vs 60-69 {b['a69']:.0f}%",
                 fontsize=11, y=1.04)
    fig.tight_layout()
    fig.savefig(out_path, dpi=150, bbox_inches="tight")
    plt.close(fig)


def main(argv=None):
    ap = argparse.ArgumentParser(description="Open-questions monitor strip.")
    ap.add_argument("--tracker",
                    default=os.path.join("logs", "claude_equity_bot_tracker.xlsx"),
                    help="the MASTER tracker (the one you paste into). "
                         "tracker_with_registry.xlsx is a GENERATED report copy.")
    ap.add_argument("--logs", default=os.path.join("logs", "signals_*.log"))
    ap.add_argument("--out-dir", default=".")
    ap.add_argument("--no-png", action="store_true")
    args = ap.parse_args(argv)

    m = collect(args.tracker, args.logs)
    md = to_markdown(m)
    print(md)
    md_path = os.path.join(args.out_dir, "monitor_strip.md")
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md + "\n")
    print(f"\nWrote: {md_path}")
    if not args.no_png:
        try:
            png = os.path.join(args.out_dir, "monitor_strip.png")
            make_png(m, png)
            print(f"Wrote: {png}")
        except Exception as e:
            print(f"(png skipped: {e})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
