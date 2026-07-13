#!/usr/bin/env python3
# =============================================================================
# blend_eval.py  —  WS1 Gate-G2 measurement (base vs blend, from the logs)
# =============================================================================
# Reads confidence_blend (+ base confidence, signal, sector/regime scores) from
# the Signal JSON log lines, joins to the tracker's +5d outcomes (computed from
# the Price columns, so it's independent of the Return-column formula cache),
# and reports what the shadow blend actually changes.
#
# WHY THE NAIVE "base-vs-blend accuracy" IS NULL (read this):
#   The blend PRESERVES DIRECTION and does NOT drive the gate (shadow-parallel).
#   So the emitted signal, its +5d outcome, and its Right? flag are IDENTICAL
#   for base and blend. Comparing base-vs-blend accuracy therefore compares a
#   thing to itself. What the blend changes is CONFIDENCE. So G2's operational
#   metric is refined to two T3-safe reads:
#
#   (A) OVERLAY PREDICTIVE VALUE  [primary] — among directional (BUY/SELL)
#       signals, do the ones the overlays CONFIRMED (blend > base) outperform
#       the ones they DISCOUNTED (blend < base)? This tests the sector/regime
#       overlay's information, NOT confidence-band ordering (T3-compliant).
#   (B) GATE-FLIP @ 70% BUY [secondary] — cases where base and blend disagree
#       on the 70% BUY threshold; would the blend's decision have been better?
#
#   Also note DP-W3 (HOLD unchanged) + a HOLD-heavy watchlist => the blend only
#   adjusts BUY/SELL, so its footprint is small; expect sparse n and read the
#   table, not a p-value. If footprint is too small after N cohorts, revisit
#   DP-W3 (whether overlays may nudge HOLD).
#
# USAGE:
#   python blend_eval.py --since 2026-07-11 --tracker tracker_with_registry.xlsx
#   python blend_eval.py --since 2026-07-11 --logs "logs\signals_*.log"
# =============================================================================

import argparse, datetime as dt, glob, json, os, re
import openpyxl

HOLD_BAND = 0.03          # +/-3% -> a HOLD was "right"
BUY_GATE  = 70            # the KEEP-70% BUY threshold


# ---------- logs: base + blend per (date, ticker) ----------
def load_log_signals(pattern):
    rows = {}
    for fp in sorted(glob.glob(pattern)):
        stem = os.path.basename(fp)
        m = re.search(r'signals_(\d{4}-\d{2}-\d{2})', stem)
        if not m:                                   # skips *_run1_coldstart etc.
            continue
        d = dt.date.fromisoformat(m.group(1))
        for line in open(fp, encoding="utf-8", errors="replace"):
            if "Signal JSON" not in line:
                continue
            j = line.split("Signal JSON:", 1)[1]
            j = j[:j.rfind("} |") + 1] if "} |" in j else j
            try:
                p = json.loads(j)
            except Exception:
                continue
            if "confidence_blend" not in p:         # pre-WS1 line -> no blend
                continue
            key = (d, str(p.get("ticker", "")).upper())
            rows[key] = dict(                        # first (AM) line per key wins
                date=d, ticker=key[1], signal=str(p.get("signal", "")).upper(),
                base=p.get("confidence"), blend=p.get("confidence_blend"),
                sector=p.get("sector_score"), regime=p.get("regime_score"),
            ) if key not in rows else rows[key]
    return rows


# ---------- tracker: +5d outcome per (date, ticker) ----------
def load_outcomes(path):
    wb = openpyxl.load_workbook(path, data_only=True); ws = wb["Signals"]
    def d(v): return v.date() if isinstance(v, dt.datetime) else (v if isinstance(v, dt.date) else None)
    def num(v): return v if isinstance(v, (int, float)) else None
    out = {}; blank = 0
    for r in range(5, ws.max_row + 1):
        dd = d(ws.cell(r, 1).value); tk = ws.cell(r, 2).value
        if dd is None and not tk:
            blank += 1
            if blank >= 30: break
            continue
        blank = 0
        if dd is None or not tk: continue
        pxs, px5 = num(ws.cell(r, 7).value), num(ws.cell(r, 8).value)
        ret5 = (px5 - pxs) / pxs if (pxs and px5) else None
        out[(dd, str(tk).strip().upper())] = ret5
    wb.close(); return out


def right(signal, ret):
    if ret is None: return None
    return abs(ret) < HOLD_BAND if signal == "HOLD" else (ret > 0 if signal == "BUY" else ret < 0)

def favorable(signal, ret):
    return ret if signal == "BUY" else (-ret if signal == "SELL" else None)


def main():
    ap = argparse.ArgumentParser(description="WS1 Gate-G2 base-vs-blend measurement.")
    ap.add_argument("--tracker", default="tracker_with_registry.xlsx")
    ap.add_argument("--logs", default=os.path.join("logs", "signals_*.log"))
    ap.add_argument("--since", default=None, help="WS1 deploy date YYYY-MM-DD (post-deploy only)")
    args = ap.parse_args()

    sig = load_log_signals(args.logs)
    out = load_outcomes(args.tracker)
    since = dt.date.fromisoformat(args.since) if args.since else None

    rows = []
    for key, s in sig.items():
        if since and s["date"] < since: continue
        s = dict(s); s["ret5"] = out.get(key)
        rows.append(s)
    rows.sort(key=lambda x: (x["date"], x["ticker"]))

    cohorts = sorted({r["date"] for r in rows})
    resolved = [r for r in rows if r["ret5"] is not None]
    adjusted = [r for r in rows if isinstance(r["base"], (int, float))
                and isinstance(r["blend"], (int, float)) and r["blend"] != r["base"]]
    print(f"WS1 blend eval  (post-deploy since {since or 'ALL'})")
    print(f"  cohorts={len(cohorts)}  signals={len(rows)}  +5d-resolved={len(resolved)}")
    print(f"  blend-adjusted={len(adjusted)}  "
          f"(lift={sum(1 for r in adjusted if r['blend']>r['base'])}, "
          f"cut={sum(1 for r in adjusted if r['blend']<r['base'])})  "
          f"[HOLD is never adjusted by design]")

    # ---------- (A) OVERLAY PREDICTIVE VALUE (T3-safe) ----------
    direc = [r for r in resolved if r["signal"] in ("BUY", "SELL")
             and isinstance(r["base"], (int, float)) and isinstance(r["blend"], (int, float))]
    confirm = [r for r in direc if r["blend"] > r["base"]]
    discount = [r for r in direc if r["blend"] < r["base"]]
    def acc(group):
        h = [r for r in group if right(r["signal"], r["ret5"])]
        fav = [favorable(r["signal"], r["ret5"]) for r in group]
        fav = [f for f in fav if f is not None]
        return (len(group), 100*len(h)/len(group) if group else 0.0,
                sum(fav)/len(fav)*100 if fav else 0.0)
    print("\n  (A) Overlay predictive value  [primary G2 read]")
    for label, g in (("overlays CONFIRMED (blend>base)", confirm),
                     ("overlays DISCOUNTED (blend<base)", discount)):
        n, a, f = acc(g)
        print(f"      {label:<34} n={n:>3}  acc={a:5.1f}%  mean favorable {f:+5.2f}%")
    if confirm and discount:
        gap = acc(confirm)[1] - acc(discount)[1]
        print(f"      -> confirm minus discount accuracy gap: {gap:+.1f}pt  "
              f"({'blend adds value' if gap > 0 else 'no separation'})")

    # ---------- (B) GATE-FLIP @ 70% BUY ----------
    buys = [r for r in resolved if r["signal"] == "BUY"
            and isinstance(r["base"], (int, float)) and isinstance(r["blend"], (int, float))]
    flips = [r for r in buys if (r["base"] >= BUY_GATE) != (r["blend"] >= BUY_GATE)]
    print(f"\n  (B) Gate-flip @ {BUY_GATE}% BUY: {len(flips)} disagreement(s) of {len(buys)} resolved BUY(s)")
    net = 0
    for r in flips:
        blend_buys = r["blend"] >= BUY_GATE          # blend would approve
        good = (r["ret5"] > 0) == blend_buys          # approve+up OR reject+down = blend right
        net += 1 if good else -1
        print(f"      {r['date']} {r['ticker']}: base {r['base']} / blend {r['blend']} "
              f"-> blend {'APPROVES' if blend_buys else 'REJECTS'}; +5d {r['ret5']*100:+.1f}%  "
              f"[{'blend better' if good else 'blend worse'}]")
    if flips:
        print(f"      -> gate-flip net: {net:+d}  ({'blend improves gate' if net > 0 else 'no improvement'})")

    # ---------- cross-validation on (A) ----------
    if len(cohorts) >= 3 and confirm and discount:
        print("\n  Cross-validation (leave-one-cohort-out, confirm-minus-discount acc gap):")
        gaps = []
        for c in cohorts:
            cf = [r for r in confirm if r["date"] != c]
            dc = [r for r in discount if r["date"] != c]
            if cf and dc:
                gaps.append(acc(cf)[1] - acc(dc)[1])
        if gaps:
            print(f"      gaps: [{', '.join(f'{g:+.0f}' for g in gaps)}]  "
                  f"min {min(gaps):+.0f}pt  (robust if all same sign)")

    print("\n  G2 verdict inputs: (A) confirm>discount robustly across folds AND "
          "(B) gate-flip net >= 0, over >= 5 cohorts. Table read, not a p-value (sparse n).")


if __name__ == "__main__":
    main()
